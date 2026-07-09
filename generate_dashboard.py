#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
素材数据仪表盘生成脚本
读取 Excel 数据源，生成交互式 HTML 仪表盘
每次更新数据后运行本脚本即可重新生成仪表盘

数据源: D:/yyzx_data/素材数据源.xlsx
输出:   D:/yyzx_data/创意周会/index.html
"""

import openpyxl
import json
import calendar
import html as html_lib
from datetime import date, datetime, timedelta
import os
import re
import importlib.util
from collections import defaultdict

# ============================================================
# 2026年中国法定节假日 & 调休
# ============================================================
# 2026年中国法定节假日 & 调休（来源：国务院办公厅通知 + chinese_calendar库）
HOLIDAYS_2026 = {
    # 元旦：1/1(周四)-1/3(周六)放假，1/4(周日)调休上班
    '2026-01-01': 'holiday', '2026-01-02': 'holiday',
    '2026-01-04': 'workday',
    # 春节：2/15(周日)-2/23(周一)放假，2/14(周六)调休上班，2/28(周六)调休上班
    '2026-02-14': 'workday',
    '2026-02-16': 'holiday', '2026-02-17': 'holiday', '2026-02-18': 'holiday',
    '2026-02-19': 'holiday', '2026-02-20': 'holiday', '2026-02-23': 'holiday',
    '2026-02-28': 'workday',
    # 清明：4/6(周一)放假（4/4-4/5为周末）
    '2026-04-06': 'holiday',
    # 劳动节：5/1(周五)-5/5(周二)放假，5/9(周六)调休上班
    '2026-05-01': 'holiday', '2026-05-04': 'holiday', '2026-05-05': 'holiday',
    '2026-05-09': 'workday',
    # 端午：6/19(周五)-6/21(周日)放假
    '2026-06-19': 'holiday',
    # 中秋：9/25(周五)-9/27(周日)放假，9/20(周日)调休上班
    '2026-09-20': 'workday', '2026-09-25': 'holiday',
    # 国庆：10/1(周四)-10/7(周三)放假，10/10(周六)调休上班
    '2026-10-01': 'holiday', '2026-10-02': 'holiday',
    '2026-10-05': 'holiday', '2026-10-06': 'holiday', '2026-10-07': 'holiday',
    '2026-10-10': 'workday',
}

HOLIDAYS_BY_YEAR = {
    2026: HOLIDAYS_2026,
}

def is_workday(d):
    ds = d.strftime('%Y-%m-%d')
    holiday_config = HOLIDAYS_BY_YEAR.get(d.year, {})
    if ds in holiday_config:
        return holiday_config[ds] == 'workday'
    return d.weekday() < 5

def count_workdays(start_date, end_date):
    cnt = 0
    cur = start_date
    while cur <= end_date:
        if is_workday(cur):
            cnt += 1
        cur += timedelta(days=1)
    return cnt

def get_month_range(ym_str):
    y, m = int(ym_str.split('-')[0]), int(ym_str.split('-')[1])
    first = date(y, m, 1)
    last_day = calendar.monthrange(y, m)[1]
    last = date(y, m, last_day)
    return first, last

# ============================================================
# 路径设置
# ============================================================
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
BASE_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH = os.path.join(BASE_DIR, '素材数据源.xlsx')
OUTPUT_PATH = os.path.join(SCRIPT_DIR, 'index.html')
TEMPLATE_PATH = os.path.join(SCRIPT_DIR, 'template.html')
ROSTER_READER_PATH = r'C:\Users\Lly621\Desktop\py_calc\code\read_my_data.py'
MANUAL_DATA_PATH = os.path.join(BASE_DIR, '运营团队手动维护数据.xlsx')


def main():
    # ============================================================
    # 读取 Excel
    # ============================================================
    print('读取 Excel 数据...')

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # --- 读取素材数据 ---
    ws_source = wb['素材数据']
    source_headers = []
    source_rows = []
    for i, row in enumerate(ws_source.iter_rows(values_only=True)):
        if i == 0:
            source_headers = list(row)
        else:
            row_data = dict(zip(source_headers, row))
            source_rows.append(row_data)
    print(f'  主数据: {len(source_rows)} 条')

    # --- 人力成本（用同一wb读取，人均人力成本自行计算） ---
    ws_labor_cost = wb['投入产出']
    labor_cost_headers = []
    labor_cost_rows = []
    for i, row in enumerate(ws_labor_cost.iter_rows(values_only=True)):
        if i == 0:
            labor_cost_headers = list(row)
        else:
            row_data = dict(zip(labor_cost_headers, row))
            row_data['ROI'] = row_data.get('ROI', 0) or 0
            # 人均人力成本 = (人力+工位) / 总人力
            total_headcount = row_data.get('总人力', 0) or 0
            total_labor_cost = row_data.get('人力+工位', 0) or 0
            row_data['人均人力成本'] = total_labor_cost / total_headcount if total_headcount > 0 else 0
            labor_cost_rows.append(row_data)
    print(f'  投入产出: {len(labor_cost_rows)} 条')


    # --- 月度人力 ---
    ws_monthly_manpower = wb['月度人力']
    monthly_manpower_headers = []
    monthly_manpower_rows = []
    for i, row in enumerate(ws_monthly_manpower.iter_rows(values_only=True)):
        if i == 0:
            monthly_manpower_headers = list(row)
        else:
            row_data = dict(zip(monthly_manpower_headers, row))
            monthly_manpower_rows.append(row_data)
    print(f'  月度人力: {len(monthly_manpower_rows)} 条')

    # --- 周度人力 ---
    ws_weekly_manpower = wb['周度人力']
    weekly_manpower_headers = []
    weekly_manpower_rows = []
    for i, row in enumerate(ws_weekly_manpower.iter_rows(values_only=True)):
        if i == 0:
            weekly_manpower_headers = list(row)
        else:
            row_data = dict(zip(weekly_manpower_headers, row))
            weekly_manpower_rows.append(row_data)
    print(f'  周度人力: {len(weekly_manpower_rows)} 条')

    # wb 不在此处 close，后面 save 时会自动处理

    # ============================================================
    # 获取近四月 & 近四周
    # ============================================================
    all_months = sorted(set(row['统计年月'] for row in source_rows if row.get('统计年月')))

    # 周度人力 Sheet 允许保留更长历史；仪表盘展示时仍只截取最近四周
    all_week_keys = set(row.get('周数') for row in weekly_manpower_rows if row.get('周数'))
    # 按“年份 + 周数”排序，兼容 26W1 / 2026W1 / 2026-W1 等格式
    def sort_week_key(w):
        if not w:
            return (0, 0)
        match = re.search(r'(\d{2,4})\D*W(\d{1,2})', str(w).upper())
        if not match:
            return (0, 0)
        year = int(match.group(1))
        week = int(match.group(2))
        if year < 100:
            year += 2000
        return (year, week)

    def get_week_end_date(w):
        year, week = sort_week_key(w)
        if year == 0 or week == 0:
            return None
        try:
            return date.fromisocalendar(year, week, 7)
        except ValueError:
            return None

    all_week_periods = sorted(list(all_week_keys), key=sort_week_key)

    RECENT_MONTHS = all_months[-4:] if len(all_months) >= 4 else all_months
    RECENT_WEEKS = all_week_periods[-4:] if len(all_week_periods) >= 4 else all_week_periods
    labor_cost_months = sorted(set(
        str(row['年月']) for row in labor_cost_rows
        if row.get('年月') and str(row['年月']) in all_months
    ))
    LABOR_COST_MONTHS = labor_cost_months[-4:] if len(labor_cost_months) >= 4 else labor_cost_months
    latest_week_end_date = get_week_end_date(RECENT_WEEKS[-1]) if RECENT_WEEKS else None
    if latest_week_end_date:
        print(f'  月度工作日截止日: {latest_week_end_date}')
    print(f'  近四月: {RECENT_MONTHS}')
    print(f'  近四周: {RECENT_WEEKS}')
    print(f'  最近已结算成本月: {LABOR_COST_MONTHS}')

    # ============================================================
    # 构建查询索引
    # ============================================================
    print('构建查询索引...')

    ROWS_BY_MONTH = defaultdict(list)
    ROWS_BY_WEEK = defaultdict(list)
    ROWS_BY_MONTH_DEPT = defaultdict(list)
    ROWS_BY_WEEK_DEPT = defaultdict(list)

    for row in source_rows:
        month_key = row.get('统计年月')
        week_key = row.get('统计周数')
        department_name = row.get('制作部门')
        if month_key:
            ROWS_BY_MONTH[month_key].append(row)
            ROWS_BY_MONTH_DEPT[(month_key, department_name)].append(row)
        if week_key:
            ROWS_BY_WEEK[week_key].append(row)
            ROWS_BY_WEEK_DEPT[(week_key, department_name)].append(row)

    MONTHLY_MANPOWER_INDEX = {
        (row.get('年月'), row.get('部门'), row.get('岗位'), row.get('业务单元')): row
        for row in monthly_manpower_rows
    }
    WEEKLY_MANPOWER_INDEX = {
        (row.get('周数'), row.get('部门'), row.get('岗位'), row.get('业务单元')): row
        for row in weekly_manpower_rows
    }
    LABOR_COST_INDEX = {
        (row.get('年月'), row.get('制作部门')): row
        for row in labor_cost_rows
    }

    print('  查询索引构建完成')

    # 单素材人力成本 — 各部门主管人力偏移量（硬编码）
    # key: 模块名, value: {部门: 偏移量}
    LABOR_COST_OFFSET = {
        '图片': {'效果设计部': 0.5, '品牌设计部': 0.5, '合肥创意部': 0, '内容二部': 0.25},
        '混剪': {'效果设计部': 0.5, '品牌设计部': 0, '合肥创意部': 0, '内容二部': 0.25},
    }

    # ============================================================
    # 辅助函数
    # ============================================================
    def parse_source_date(value):
        if isinstance(value, str):
            try:
                return date.fromisoformat(value)
            except ValueError:
                return None
        if hasattr(value, 'year'):
            return date(value.year, value.month, value.day)
        return None

    def get_content_editing_week_key(row):
        source_date = parse_source_date(row.get('统计日期'))
        if not source_date:
            return row.get('统计周数')
        iso_year, iso_week, _ = (source_date + timedelta(days=3)).isocalendar()
        return f'{iso_year % 100:02d}W{iso_week}'

    def filter_source_rows(business_units=None, posts=None, biz_source=None,
                           exclude_data_sources=None, months=None, weeks=None, depts=None,
                           week_key_getter=None):
        candidates = None
        if months:
            candidates = []
            for month_key in months:
                if depts:
                    for dept in depts:
                        candidates.extend(ROWS_BY_MONTH_DEPT.get((month_key, dept), []))
                else:
                    candidates.extend(ROWS_BY_MONTH.get(month_key, []))
        elif weeks and not week_key_getter:
            candidates = []
            for week_key in weeks:
                if depts:
                    for dept in depts:
                        candidates.extend(ROWS_BY_WEEK_DEPT.get((week_key, dept), []))
                else:
                    candidates.extend(ROWS_BY_WEEK.get(week_key, []))
        else:
            candidates = source_rows

        filtered_rows = []
        for row in candidates:
            if exclude_data_sources and row.get('数据来源') in exclude_data_sources:
                continue
            if depts and row.get('制作部门') not in depts:
                continue
            if business_units and row.get('业务单元') not in business_units:
                continue
            if posts and row.get('岗位') not in posts:
                continue
            if biz_source and row.get('业务来源') != biz_source:
                continue
            if months and row.get('统计年月') not in months:
                continue
            row_week_key = week_key_getter(row) if week_key_getter else row.get('统计周数')
            if weeks and row_week_key not in weeks:
                continue
            filtered_rows.append(row)
        return filtered_rows

    def get_manpower_value(manpower_rows, period_key, dept, post, business_unit, field='产出用'):
        """查人力数，按月或周查找（修复：现在正确使用key_val匹配年月或周数）"""
        index = MONTHLY_MANPOWER_INDEX if manpower_rows is monthly_manpower_rows else WEEKLY_MANPOWER_INDEX
        row = index.get((period_key, dept, post, business_unit))
        if not row:
            return 0
        value = row.get(field)
        return value if value is not None else 0

    def get_labor_cost_per_capita(month_key, dept):
        """查某月某部门的人均人力成本"""
        row = LABOR_COST_INDEX.get((month_key, dept))
        if not row:
            return 0
        value = row.get('人均人力成本')
        return value if value is not None else 0

    def calc_single_labor_cost(total_output, dept_manpower_list):
        """
        计算单素材人力成本
        total_output: 总产出
        dept_manpower_list: [(产出用人力, 偏移量, 人均人力成本), ...]
        公式: Σ(产出用人力 + 偏移量) × 人均人力成本 ÷ 总产出
        """
        if total_output == 0:
            return 0
        total_cost = sum((mp + offset) * cost for mp, offset, cost in dept_manpower_list)
        return total_cost / total_output


    def sum_output(rows, output_field):
        return sum((row.get(output_field, 0) or 0) for row in rows)

    def calculate_metrics(rows, workdays, manpower_output, manpower_workload, output_field='产出数', debug_info=None,
                          efficiency_output_field=None):
        total_output = sum_output(rows, output_field)
        efficiency_output = sum_output(rows, efficiency_output_field or output_field)
        total_workload = sum((r.get('周报工作量', 0) or 0) for r in rows)

        if debug_info:
            print(f"[{debug_info['type']}] {debug_info['module']} | {debug_info['dept']} | {debug_info['period']}")
            print(f"  总产出={total_output:.1f}, 总工时={total_workload:.1f}, 工作日={workdays}")
            print(f"  产出用={manpower_output:.1f}, 工作量用={manpower_workload:.1f}")

        if workdays == 0 or manpower_output == 0:
            avg_daily_output = 0
        else:
            avg_daily_output = total_output / workdays / manpower_output

        if workdays == 0 or manpower_workload == 0:
            avg_daily_workload = 0
        else:
            avg_daily_workload = total_workload / workdays / manpower_workload

        saturation = avg_daily_workload / 7 if avg_daily_workload else 0

        if efficiency_output == 0:
            single_time = 0
        else:
            single_time = total_workload * 60 / efficiency_output

        if debug_info:
            print(f"  => 人均日均产出={avg_daily_output:.1f} (公式: {total_output:.1f}/{workdays}/{manpower_output:.1f})")
            print(f"  => 人均日均工时={avg_daily_workload:.1f} (公式: {total_workload:.1f}/{workdays}/{manpower_workload:.1f})")
            print(f"  => 饱和度={saturation:.1f} (公式: {avg_daily_workload:.1f}/7)")
            if efficiency_output > 0:
                print(f"  => 单素材耗时={single_time:.1f} (公式: {total_workload:.1f}*60/{efficiency_output:.1f})")
            print()

        return {
            'total_output': round(total_output, 2),
            'avg_daily_output': round(avg_daily_output, 4),
            'avg_daily_workload': round(avg_daily_workload, 4),
            'saturation': round(saturation, 4),
            'single_time': round(single_time, 2),
            'workdays': workdays,
            'output_manpower': round(manpower_output, 2),
            'workload_manpower': round(manpower_workload, 2),
            'total_workload': round(total_workload, 2),
            'efficiency_output': round(efficiency_output, 2),
        }


    def calculate_efficiency_metrics(rows, workdays, manpower_workload, output_field='产出数'):
        """仅计算效率指标（饱和度、人均日均工时、单素材耗时），用于排除实习生后重算"""
        total_output = sum_output(rows, output_field)
        total_workload = sum((r.get('周报工作量', 0) or 0) for r in rows)

        if workdays == 0 or manpower_workload == 0:
            avg_daily_workload = 0
        else:
            avg_daily_workload = total_workload / workdays / manpower_workload

        saturation = avg_daily_workload / 7 if avg_daily_workload else 0

        if total_output == 0:
            single_time = 0
        else:
            single_time = total_workload * 60 / total_output

        return {
            'avg_daily_workload': round(avg_daily_workload, 4),
            'saturation': round(saturation, 4),
            'single_time': round(single_time, 2),
        }

    def exclude_intern_rows(rows):
        """排除是否实习生=1的记录。"""
        return [row for row in rows if row.get('是否实习生') != 1]

    def count_distinct_projects(rows):
        """按修正所属中心+修正需求部门+修正产品去重，计算项目数"""
        projects = set()
        for r in rows:
            center = (r.get('修正所属中心') or '').strip()
            dept_req = (r.get('修正需求部门') or '').strip()
            product = (r.get('修正产品') or '').strip()
            projects.add((center, dept_req, product))
        # 过滤掉三个字段都为空的无效记录
        valid = {p for p in projects if any(p)}
        return len(valid)

    PERSON_EFFICIENCY_CLASSES = ['高效率', '正常', '待提升']
    PERSON_EFFICIENCY_MONTHS = ['2026-03', '2026-04', '2026-05']
    PERSON_EFFICIENCY_PERIOD = '2026-03~2026-05'
    PERSON_HIGH_EFFICIENCY_THRESHOLD = 1.2
    PERSON_NORMAL_EFFICIENCY_THRESHOLD = 1.0
    PERSON_STANDARD_CARD_ORDER = {
        '混剪': ['常规混剪', '复杂混剪', '混剪修改', '脚本&剪辑'],
        '图片': ['整图设计', '套模板整图', '复杂修改', '简单修改', '宫格小图', '轮播小图', '设计找图', '多屏长图', '落地页设计', '创意品牌图'],
        '实拍': ['实拍剪辑'],
    }

    def pe_norm(value):
        return str(value).strip() if value is not None else ''

    def pe_number(value):
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    def pe_settlement_dept(row):
        dept = pe_norm(row.get('制作部门-结算')) or pe_norm(row.get('制作部门'))
        if dept == '品牌设计部-内容':
            return '内容一部'
        return dept

    def pe_page_business_unit(value):
        unit = pe_norm(value)
        if unit == '内容部常规':
            return '实拍'
        return unit

    def pe_page_demand_type(row):
        if pe_norm(row.get('业务单元')) == '内容部常规':
            return '实拍剪辑'
        return pe_norm(row.get('需求类型'))

    def pe_standard_key_for_row(row):
        return (
            pe_page_business_unit(row.get('业务单元')),
            pe_page_demand_type(row),
        )

    def pe_classify_efficiency(value):
        if value is None:
            return '无标准'
        if value >= PERSON_HIGH_EFFICIENCY_THRESHOLD:
            return '高效率'
        if value >= PERSON_NORMAL_EFFICIENCY_THRESHOLD:
            return '正常'
        return '待提升'

    def pe_parse_date(value):
        if value is None or value == '':
            return None
        if isinstance(value, datetime):
            return date(value.year, value.month, value.day)
        if isinstance(value, date):
            return value
        if hasattr(value, 'to_pydatetime'):
            value = value.to_pydatetime()
        if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
            return date(value.year, value.month, value.day)
        text = pe_norm(value)
        if not text or text.lower() in {'nat', 'nan', 'none'}:
            return None
        for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y-%m-%d %H:%M:%S'):
            try:
                parsed = datetime.strptime(text[:19], fmt)
                return date(parsed.year, parsed.month, parsed.day)
            except ValueError:
                pass
        return None

    def pe_format_date(value):
        parsed = pe_parse_date(value)
        return parsed.isoformat() if parsed else ''

    def read_employee_roster(cutoff_date):
        """读取花名册，用制作人员姓名匹配员工状态和在职年数。"""
        if not os.path.exists(ROSTER_READER_PATH):
            print(f'  花名册读取跳过: 未找到 {ROSTER_READER_PATH}')
            return {}
        try:
            spec = importlib.util.spec_from_file_location('external_roster_reader', ROSTER_READER_PATH)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            roster_df = module.read_my_data('SELECT `姓名`, `员工状态`, `入职时间`, `离职时间` FROM `花名册`')
        except Exception as exc:
            print(f'  花名册读取失败: {exc}')
            return {}

        roster = {}
        duplicate_names = set()
        for _, row in roster_df.iterrows():
            name = pe_norm(row.get('姓名'))
            if not name:
                continue
            status = pe_norm(row.get('员工状态'))
            join_date = pe_parse_date(row.get('入职时间'))
            leave_date = pe_parse_date(row.get('离职时间'))
            is_resigned = ('离职' in status) or (leave_date is not None and leave_date <= cutoff_date)
            end_date = leave_date if is_resigned and leave_date else cutoff_date
            tenure_years = None
            if join_date and end_date >= join_date:
                tenure_years = round((end_date - join_date).days / 365.25, 1)
            item = {
                'employeeStatus': status or ('已离职' if is_resigned else '在职'),
                'isResigned': is_resigned,
                'joinDate': join_date.isoformat() if join_date else '',
                'leaveDate': leave_date.isoformat() if leave_date else '',
                'tenureYears': tenure_years,
                'rosterMatched': True,
            }
            if name in roster:
                duplicate_names.add(name)
                old = roster[name]
                old_active = not old.get('isResigned')
                new_active = not item.get('isResigned')
                if old_active and not new_active:
                    continue
                if old_active == new_active and (old.get('joinDate') or '') >= (item.get('joinDate') or ''):
                    continue
            roster[name] = item

        print(f'  花名册: {len(roster)} 人')
        if duplicate_names:
            print(f'  花名册同名人员: {len(duplicate_names)} 人，已优先保留在职或较新入职记录')
        return roster

    def is_person_efficiency_row(row):
        business_unit = pe_norm(row.get('业务单元'))
        post = pe_norm(row.get('岗位'))
        return (
            pe_norm(row.get('业务来源')) == '广告投流'
            and (
                (business_unit == '图片' and post == '设计')
                or (business_unit == '混剪' and post == '剪辑')
                or (business_unit == '内容部常规' and post == '剪辑')
            )
            and pe_number(row.get('产出数')) > 0
            and pe_number(row.get('周报工作量')) > 0
            and pe_number(row.get('是否实习生')) != 1
            and pe_norm(row.get('制作人员'))
            and pe_norm(row.get('统计年月')) in PERSON_EFFICIENCY_MONTHS
        )

    def read_material_efficiency_standards():
        standard_values = defaultdict(list)
        if not os.path.exists(MANUAL_DATA_PATH):
            print(f'  素材效率标准读取跳过: 未找到 {MANUAL_DATA_PATH}')
            return {}, []
        try:
            wb_manual = openpyxl.load_workbook(MANUAL_DATA_PATH, read_only=True, data_only=True)
            ws_standard = wb_manual['素材效率标准']
            headers = [pe_norm(value) for value in next(ws_standard.iter_rows(values_only=True))]
            idx = {name: i for i, name in enumerate(headers)}
            required = ['业务单元', '需求类型', '标准耗时']
            missing = [name for name in required if name not in idx]
            if missing:
                print(f"  素材效率标准读取失败: 缺少字段 {', '.join(missing)}")
                wb_manual.close()
                return {}, []
            for row in ws_standard.iter_rows(min_row=2, values_only=True):
                unit = pe_page_business_unit(row[idx['业务单元']])
                demand_type = pe_norm(row[idx['需求类型']])
                standard_minutes = pe_number(row[idx['标准耗时']])
                if not unit or not demand_type or standard_minutes <= 0:
                    continue
                standard_values[(unit, demand_type)].append(standard_minutes)
            wb_manual.close()
        except Exception as exc:
            print(f'  素材效率标准读取失败: {exc}')
            return {}, []
        standards = {}
        standard_cards = defaultdict(list)
        for (unit, demand_type), values in standard_values.items():
            standard_minutes = sum(values) / len(values)
            standards[(unit, demand_type)] = {
                'businessUnit': unit,
                'demandType': demand_type,
                'standardMinutes': standard_minutes,
                'standardHours': standard_minutes / 60,
                'sourceCount': len(values),
            }
            standard_cards[unit].append({
                'businessUnit': unit,
                'demandType': demand_type,
                'standardMinutes': round(standard_minutes, 2),
                'sourceCount': len(values),
            })
        def card_sort_key(unit, item):
            order = PERSON_STANDARD_CARD_ORDER.get(unit, [])
            try:
                index = order.index(item['demandType'])
            except ValueError:
                index = len(order) + 999
            return (index, item['demandType'])

        standard_cards = {
            unit: sorted(items, key=lambda item, unit_name=unit: card_sort_key(unit_name, item))
            for unit, items in sorted(standard_cards.items(), key=lambda pair: pair[0])
        }
        print(f'  素材效率标准: {len(standards)} 条')
        return standards, standard_cards

    PERSON_ROSTER_BY_NAME = read_employee_roster(date(2026, 5, 31))
    PERSON_STANDARD_BY_TYPE, PERSON_STANDARD_CARDS = read_material_efficiency_standards()

    def pe_roster_info(person):
        return PERSON_ROSTER_BY_NAME.get(person, {
            'employeeStatus': '未匹配',
            'isResigned': False,
            'joinDate': '',
            'leaveDate': '',
            'tenureYears': None,
            'tenureDays': None,
            'rosterMatched': False,
        })

    def pe_tenure_days(roster_info):
        join_date = pe_parse_date(roster_info.get('joinDate'))
        leave_date = pe_parse_date(roster_info.get('leaveDate'))
        if not join_date:
            return None
        end_date = leave_date if leave_date else date(2026, 5, 31)
        if end_date < join_date:
            return None
        return (end_date - join_date).days + 1

    def pe_standard_for_row(row):
        return PERSON_STANDARD_BY_TYPE.get(pe_standard_key_for_row(row))

    def new_person_efficiency_item(period, row, business_unit):
        person = pe_norm(row.get('制作人员'))
        roster_info = pe_roster_info(person)
        tenure_days = pe_tenure_days(roster_info)
        return {
            'period': period,
            'dept': pe_settlement_dept(row),
            'person': person,
            'employeeStatus': roster_info['employeeStatus'],
            'isResigned': roster_info['isResigned'],
            'joinDate': roster_info['joinDate'],
            'leaveDate': roster_info['leaveDate'],
            'tenureYears': roster_info['tenureYears'],
            'tenureDays': tenure_days,
            'rosterMatched': roster_info['rosterMatched'],
            'businessUnit': business_unit,
            'totalOutput': 0.0,
            'totalWorkload': 0.0,
            'evaluatedOutput': 0.0,
            'evaluatedWorkload': 0.0,
            'standardWorkload': 0.0,
            'sampleMissingOutput': 0.0,
            'sampleMissingWorkload': 0.0,
            'monthly': {month: {'standardWorkload': 0.0, 'evaluatedWorkload': 0.0} for month in PERSON_EFFICIENCY_MONTHS},
        }

    def finish_person_efficiency_item(item):
        standard = item['standardWorkload']
        total_output = item['totalOutput']
        evaluated_output = item['evaluatedOutput']
        evaluated_workload = item['evaluatedWorkload']
        if standard > 0 and evaluated_workload > 0:
            item['efficiency'] = round(standard / evaluated_workload, 4)
        else:
            item['efficiency'] = None
        item['standardCoverage'] = round(evaluated_output / total_output, 4) if total_output else 0
        item['sampleMissingShare'] = round(item['sampleMissingOutput'] / total_output, 4) if total_output else 0
        item['eligible'] = standard > 0 and evaluated_workload > 0
        item['className'] = pe_classify_efficiency(item['efficiency'])
        item['monthlyEfficiency'] = {}
        for month, month_item in item['monthly'].items():
            if month_item['standardWorkload'] > 0 and month_item['evaluatedWorkload'] > 0:
                item['monthlyEfficiency'][month] = round(month_item['standardWorkload'] / month_item['evaluatedWorkload'], 4)
            else:
                item['monthlyEfficiency'][month] = None
        for key in ['totalOutput', 'totalWorkload', 'evaluatedOutput', 'evaluatedWorkload',
                    'standardWorkload', 'sampleMissingOutput', 'sampleMissingWorkload']:
            item[key] = round(item[key], 2)
        item.pop('monthly', None)
        return item

    def summarize_person_efficiency_period(period, rows):
        person_items = {}
        detail_items = {}
        project_detail_items = {}
        project_items = {}

        for row in rows:
            month = pe_norm(row.get('统计年月'))
            output = pe_number(row.get('产出数'))
            workload = pe_number(row.get('周报工作量'))
            estimated_fee = pe_number(row.get('预估费用'))
            standard = pe_standard_for_row(row)
            standard_workload = output * standard['standardHours'] if standard else 0.0

            person_keys = [
                (pe_settlement_dept(row), pe_norm(row.get('制作人员')), pe_page_business_unit(row.get('业务单元'))),
                (pe_settlement_dept(row), pe_norm(row.get('制作人员')), '总计'),
            ]
            for person_key in person_keys:
                item = person_items.setdefault(person_key, new_person_efficiency_item(period, row, person_key[2]))
                item['totalOutput'] += output
                item['totalWorkload'] += workload
                if standard:
                    item['evaluatedOutput'] += output
                    item['evaluatedWorkload'] += workload
                    item['standardWorkload'] += standard_workload
                    if month in item['monthly']:
                        item['monthly'][month]['standardWorkload'] += standard_workload
                        item['monthly'][month]['evaluatedWorkload'] += workload
                else:
                    item['sampleMissingOutput'] += output
                    item['sampleMissingWorkload'] += workload

            detail_key = (
                pe_settlement_dept(row),
                pe_norm(row.get('制作人员')),
                pe_page_business_unit(row.get('业务单元')),
                pe_norm(row.get('修正需求部门')),
                pe_norm(row.get('修正产品')),
                pe_page_demand_type(row),
            )
            detail = detail_items.setdefault(detail_key, {
                'period': period,
                'dept': detail_key[0],
                'person': detail_key[1],
                'businessUnit': detail_key[2],
                'demandDept': detail_key[3],
                'product': detail_key[4],
                'demandType': detail_key[5],
                'output': 0.0,
                'workload': 0.0,
                'estimatedFee': 0.0,
                'evaluatedOutput': 0.0,
                'evaluatedWorkload': 0.0,
                'standardWorkload': 0.0,
                'standardMinutes': standard['standardMinutes'] if standard else None,
                'standardMatched': bool(standard),
                'monthly': {month: {'standardWorkload': 0.0, 'evaluatedWorkload': 0.0} for month in PERSON_EFFICIENCY_MONTHS},
            })
            detail['output'] += output
            detail['workload'] += workload
            detail['estimatedFee'] += estimated_fee
            if standard:
                detail['evaluatedOutput'] += output
                detail['evaluatedWorkload'] += workload
                detail['standardWorkload'] += standard_workload
                if month in detail['monthly']:
                    detail['monthly'][month]['standardWorkload'] += standard_workload
                    detail['monthly'][month]['evaluatedWorkload'] += workload

            project_detail_key = (
                pe_settlement_dept(row),
                pe_page_business_unit(row.get('业务单元')),
                pe_norm(row.get('修正需求部门')),
                pe_norm(row.get('修正产品')),
                pe_page_demand_type(row),
            )
            project_detail = project_detail_items.setdefault(project_detail_key, {
                'period': period,
                'dept': project_detail_key[0],
                'businessUnit': project_detail_key[1],
                'demandDept': project_detail_key[2],
                'product': project_detail_key[3],
                'demandType': project_detail_key[4],
                'output': 0.0,
                'workload': 0.0,
                'estimatedFee': 0.0,
                'evaluatedOutput': 0.0,
                'evaluatedWorkload': 0.0,
                'standardWorkload': 0.0,
                'monthly': {month: {'standardWorkload': 0.0, 'evaluatedWorkload': 0.0} for month in PERSON_EFFICIENCY_MONTHS},
            })
            project_detail['output'] += output
            project_detail['workload'] += workload
            project_detail['estimatedFee'] += estimated_fee
            if standard:
                project_detail['evaluatedOutput'] += output
                project_detail['evaluatedWorkload'] += workload
                project_detail['standardWorkload'] += standard_workload
                if month in project_detail['monthly']:
                    project_detail['monthly'][month]['standardWorkload'] += standard_workload
                    project_detail['monthly'][month]['evaluatedWorkload'] += workload

            project_key = (
                pe_norm(row.get('修正所属中心')),
                pe_norm(row.get('修正需求部门')),
                pe_norm(row.get('修正产品')),
                pe_page_business_unit(row.get('业务单元')),
            )
            project = project_items.setdefault(project_key, {
                'period': period,
                'center': project_key[0],
                'demandDept': project_key[1],
                'product': project_key[2],
                'businessUnit': project_key[3],
                'output': 0.0,
                'workload': 0.0,
                'evaluatedOutput': 0.0,
                'evaluatedWorkload': 0.0,
                'standardWorkload': 0.0,
                'depts': set(),
                'persons': set(),
                'people': {},
            })
            project['output'] += output
            project['workload'] += workload
            project['depts'].add(pe_settlement_dept(row))
            project['persons'].add(pe_norm(row.get('制作人员')))
            if standard:
                project['evaluatedOutput'] += output
                project['evaluatedWorkload'] += workload
                project['standardWorkload'] += standard_workload

            project_person_key = (pe_settlement_dept(row), pe_norm(row.get('制作人员')))
            project_roster_info = pe_roster_info(project_person_key[1])
            project_person = project['people'].setdefault(project_person_key, {
                'dept': project_person_key[0],
                'person': project_person_key[1],
                'employeeStatus': project_roster_info['employeeStatus'],
                'isResigned': project_roster_info['isResigned'],
                'joinDate': project_roster_info['joinDate'],
                'leaveDate': project_roster_info['leaveDate'],
                'tenureYears': project_roster_info['tenureYears'],
                'tenureDays': pe_tenure_days(project_roster_info),
                'rosterMatched': project_roster_info['rosterMatched'],
                'output': 0.0,
                'workload': 0.0,
                'evaluatedOutput': 0.0,
                'evaluatedWorkload': 0.0,
                'standardWorkload': 0.0,
                'sampleMissingOutput': 0.0,
            })
            project_person['output'] += output
            project_person['workload'] += workload
            if standard:
                project_person['evaluatedOutput'] += output
                project_person['evaluatedWorkload'] += workload
                project_person['standardWorkload'] += standard_workload
            else:
                project_person['sampleMissingOutput'] += output

        rows_result = [finish_person_efficiency_item(item) for item in person_items.values()]

        details_result = []
        for detail in detail_items.values():
            standard_workload = detail['standardWorkload']
            detail['efficiency'] = round(standard_workload / detail['evaluatedWorkload'], 4) if standard_workload > 0 and detail['evaluatedWorkload'] > 0 else None
            detail['monthlyEfficiency'] = {}
            for month, month_item in detail['monthly'].items():
                if month_item['standardWorkload'] > 0 and month_item['evaluatedWorkload'] > 0:
                    detail['monthlyEfficiency'][month] = round(month_item['standardWorkload'] / month_item['evaluatedWorkload'], 4)
                else:
                    detail['monthlyEfficiency'][month] = None
            for key in ['output', 'workload', 'estimatedFee', 'evaluatedOutput', 'evaluatedWorkload', 'standardWorkload']:
                detail[key] = round(detail[key], 2)
            detail.pop('monthly', None)
            details_result.append(detail)

        project_details_result = []
        for project_detail in project_detail_items.values():
            standard_workload = project_detail['standardWorkload']
            project_detail['efficiency'] = round(standard_workload / project_detail['evaluatedWorkload'], 4) if standard_workload > 0 and project_detail['evaluatedWorkload'] > 0 else None
            project_detail['monthlyEfficiency'] = {}
            for month, month_item in project_detail['monthly'].items():
                if month_item['standardWorkload'] > 0 and month_item['evaluatedWorkload'] > 0:
                    project_detail['monthlyEfficiency'][month] = round(month_item['standardWorkload'] / month_item['evaluatedWorkload'], 4)
                else:
                    project_detail['monthlyEfficiency'][month] = None
            project_detail['singleTimeMinutes'] = round(project_detail['evaluatedWorkload'] * 60 / project_detail['evaluatedOutput'], 2) if project_detail['evaluatedOutput'] > 0 and project_detail['evaluatedWorkload'] > 0 else None
            project_detail['unitFee'] = round(project_detail['estimatedFee'] / project_detail['output'], 2) if project_detail['output'] > 0 else None
            for key in ['output', 'workload', 'estimatedFee', 'evaluatedOutput', 'evaluatedWorkload', 'standardWorkload']:
                project_detail[key] = round(project_detail[key], 2)
            project_detail.pop('monthly', None)
            project_details_result.append(project_detail)

        top_projects = []
        for project in project_items.values():
            if len(project['depts']) < 2:
                continue
            standard_workload = project['standardWorkload']
            project['deptCount'] = len(project['depts'])
            project['personCount'] = len(project['persons'])
            project['efficiency'] = round(standard_workload / project['evaluatedWorkload'], 4) if standard_workload > 0 and project['evaluatedWorkload'] > 0 else None
            project['standardCoverage'] = round(project['evaluatedOutput'] / project['output'], 4) if project['output'] else 0
            class_counts = {name: 0 for name in PERSON_EFFICIENCY_CLASSES}
            people = []
            for person_item in project['people'].values():
                person_standard = person_item['standardWorkload']
                person_output = person_item['output']
                evaluated_output = person_item['evaluatedOutput']
                person_item['efficiency'] = round(person_standard / person_item['evaluatedWorkload'], 4) if person_standard > 0 and person_item['evaluatedWorkload'] > 0 else None
                person_item['standardCoverage'] = round(evaluated_output / person_output, 4) if person_output else 0
                person_item['sampleMissingShare'] = round(person_item['sampleMissingOutput'] / person_output, 4) if person_output else 0
                person_item['className'] = pe_classify_efficiency(person_item['efficiency']) if person_standard > 0 else '无标准'
                if person_item['className'] in class_counts:
                    class_counts[person_item['className']] += 1
                for key in ['output', 'workload', 'evaluatedOutput', 'evaluatedWorkload',
                            'standardWorkload', 'sampleMissingOutput']:
                    person_item[key] = round(person_item[key], 2)
                people.append(person_item)
            project['people'] = sorted(
                people,
                key=lambda x: x['efficiency'] if x['efficiency'] is not None else -999,
                reverse=True
            )
            project['classCounts'] = class_counts
            for key in ['output', 'workload', 'evaluatedOutput', 'evaluatedWorkload', 'standardWorkload']:
                project[key] = round(project[key], 2)
            project['depts'] = sorted(project['depts'])
            project['persons'] = sorted(project['persons'])
            top_projects.append(project)

        top_projects = sorted(top_projects, key=lambda x: x['output'], reverse=True)[:10]
        return rows_result, details_result, project_details_result, top_projects

    def build_person_efficiency_data():
        base_rows = [row for row in source_rows if is_person_efficiency_row(row)]
        periods = [PERSON_EFFICIENCY_PERIOD]
        all_rows = []
        all_details = []
        top_projects_by_period = {}

        person_rows, detail_rows, project_detail_rows, top_projects = summarize_person_efficiency_period(PERSON_EFFICIENCY_PERIOD, base_rows)
        all_rows.extend(person_rows)
        all_details.extend(detail_rows)
        all_project_details = project_detail_rows
        top_projects_by_period[PERSON_EFFICIENCY_PERIOD] = top_projects

        departments = sorted({row['dept'] for row in all_rows if row.get('dept')})
        department_summary = []
        for dept in departments:
            dept_rows = [
                row for row in all_rows
                if row.get('dept') == dept and row.get('businessUnit') == '总计' and row.get('eligible')
            ]
            if not dept_rows:
                continue
            class_counts = {name: 0 for name in PERSON_EFFICIENCY_CLASSES}
            total_standard = 0.0
            total_actual = 0.0
            for row in dept_rows:
                total_standard += row.get('standardWorkload', 0) or 0
                total_actual += row.get('evaluatedWorkload', 0) or 0
                if row.get('className') in class_counts:
                    class_counts[row['className']] += 1
            department_summary.append({
                'dept': dept,
                'efficiency': round(total_standard / total_actual, 4) if total_standard > 0 and total_actual > 0 else None,
                'eligibleCount': len(dept_rows),
                'highCount': class_counts['高效率'],
                'normalCount': class_counts['正常'],
                'watchCount': class_counts['待提升'],
            })
        return {
            'periods': periods,
            'businessUnits': ['总计', '图片', '混剪', '实拍'],
            'departments': departments,
            'classNames': PERSON_EFFICIENCY_CLASSES,
            'months': PERSON_EFFICIENCY_MONTHS,
            'rows': all_rows,
            'details': all_details,
            'projectDetails': all_project_details,
            'departmentSummary': department_summary,
            'topProjectsByPeriod': top_projects_by_period,
            'standardCards': PERSON_STANDARD_CARDS,
            'readme': {
                'purpose': [
                    '用于评估广告投流业务中创意人员在标准耗时口径下的效率表现。',
                ],
                'metric': [
                    {
                        'title': '报表统计范围',
                        'text': '',
                        'highlight': '【26M3~26M5】',
                    },
                    {
                        'title': '实际耗时',
                        'text': '各团队每月提报的各项目实际制作耗时',
                    },
                    {
                        'title': '标准耗时',
                        'text': '基于上半年实际制作数据沉淀的类型平均水准，用于标准化效率评估口径（详见右表）',
                    },
                    {
                        'title': '评估效率',
                        'text': '标准耗时 / 实际耗时，数值越高表示效率越快',
                        'childrenTitle': '效率分类规则',
                        'children': [
                            '高效率：评估效率 >= 1.2',
                            '正常：1.0 <= 评估效率 < 1.2',
                            '待提升：评估效率 < 1.0',
                        ],
                    },
                    {
                        'title': '口径提示',
                        'text': '本表聚焦标准化口径下的产出效率，少量未纳入标准的素材类型已剔除（如海报，视频制作，图片其他，单图制作等），故不代表员工完整产出量或全部工作量。',
                    },
                ],
            },
        }

    # 计算每月工作日
    month_workdays = {}
    for month_key in sorted(set(RECENT_MONTHS + LABOR_COST_MONTHS)):
        first, last = get_month_range(month_key)
        if latest_week_end_date:
            last = min(last, latest_week_end_date)
        month_workdays[month_key] = count_workdays(first, last) if first <= last else 0

    # 计算每周工作日
    week_workdays = {}
    for w in RECENT_WEEKS:
        dates_in_week = set()
        for row in source_rows:
            if row.get('统计周数') == w and row.get('统计日期'):
                ds = row['统计日期']
                if isinstance(ds, str):
                    try:
                        dates_in_week.add(date.fromisoformat(ds))
                    except:
                        pass
                elif hasattr(ds, 'year'):
                    dates_in_week.add(date(ds.year, ds.month, ds.day))
        if dates_in_week:
            week_workdays[w] = count_workdays(min(dates_in_week), max(dates_in_week))
        else:
            week_workdays[w] = 5

    print(f'  每月工作日: {month_workdays}')
    print(f'  每周工作日: {week_workdays}')

    # ============================================================
    # 模块定义
    # ============================================================
    # `MODULES` 是新增模块时最主要的入口，推荐优先在这里描述差异，不去主流程里加分支。
    # 字段约定：
    # - `name`: 仪表盘上展示的模块名称，也是某些专项规则（如单素材人力成本偏移量）的关联键。
    # - `filters`: 定义源数据怎么筛。`business_units` / `posts` / `biz_source` 都是等值筛选；
    #   `exclude_data_sources` 用来显式声明哪些数据来源要排除。
    # - `manpower`: 定义人力口径。月度/周度可能走不同业务单元，`department_position_map`
    #   用来把“部门 -> 岗位”对齐到人力表；`department_scope=None` 表示按产出自动排序部门。
    # - `metrics`: 定义指标规则。`output_field` 决定产出取哪一列；
    #   `enable_project_count` / `exclude_intern_efficiency` / `show_total` / `hidden`
    #   都是显式规则，避免靠模块名或字段名做隐式推断。
    # 新增模块时的建议顺序：
    # 1. 先确认要筛哪些源数据，补 `filters`
    # 2. 再确认人力表该怎么对齐，补 `manpower`
    # 3. 最后确认产出字段、项目数、总计展示等规则，补 `metrics`
    MODULES = [
        {
            'name': '图片',
            'filters': {
                'business_units': ['图片'],
                'posts': ['设计'],
                'biz_source': None,
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '图片',
                'weekly_business_unit': '图片',
                'department_scope': None,
                'department_position_map': {'效果设计部': '设计', '品牌设计部': '设计',
                                            '合肥创意部': '设计', '内容二部': '设计'},
            },
            'metrics': {
                'output_field': '产出数',
                'enable_project_count': False,
                'exclude_intern_efficiency': True,
                'show_total': True,
                'hidden': False,
            },
        },
        {
            'name': '混剪',
            'filters': {
                'business_units': ['混剪'],
                'posts': ['剪辑'],
                'biz_source': None,
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '混剪',
                'weekly_business_unit': '混剪',
                'department_scope': None,
                'department_position_map': {'效果设计部': '剪辑', '品牌设计部': '剪辑',
                                            '合肥创意部': '剪辑', '内容二部': '剪辑'},
            },
            'metrics': {
                'output_field': '产出数',
                'enable_project_count': False,
                'exclude_intern_efficiency': True,
                'show_total': True,
                'hidden': False,
            },
        },
        {
            'name': '内容团队-编剧',
            'filters': {
                'business_units': ['混剪', '内容部常规'],
                'posts': ['编剧'],
                'biz_source': '广告投流',
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '拍摄',
                'weekly_business_unit': '拍摄',
                'department_scope': ['品牌设计部', '内容二部'],
                'department_position_map': {'品牌设计部': '编剧', '内容二部': '编剧'},
            },
            'metrics': {
                'output_field': '写脚本数',
                'alternate_output_field': '产出数',
                'efficiency_output_field': '产出数',
                'enable_project_count': True,
                'exclude_intern_efficiency': False,
                'show_total': False,
                'hidden': False,
            },
        },
        {
            'name': '内容团队-导演',
            'filters': {
                'business_units': ['内容部常规'],
                'posts': ['导演'],
                'biz_source': '广告投流',
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '拍摄',
                'weekly_business_unit': '拍摄',
                'department_scope': ['品牌设计部', '内容二部'],
                'department_position_map': {'品牌设计部': '导演', '内容二部': '导演'},
            },
            'metrics': {
                'output_field': '产出数',
                'enable_project_count': False,
                'exclude_intern_efficiency': False,
                'show_total': False,
                'hidden': True,
            },
        },
        {
            'name': '内容团队-摄像',
            'filters': {
                'business_units': ['内容部常规'],
                'posts': ['摄像'],
                'biz_source': '广告投流',
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '拍摄',
                'weekly_business_unit': '拍摄',
                'department_scope': ['品牌设计部', '内容二部'],
                'department_position_map': {'品牌设计部': '摄像', '内容二部': '摄像'},
            },
            'metrics': {
                'output_field': '产出数',
                'enable_project_count': False,
                'exclude_intern_efficiency': False,
                'show_total': False,
                'hidden': False,
            },
        },
        {
            'name': '内容团队-剪辑',
            'filters': {
                'business_units': ['内容部常规'],
                'posts': ['剪辑'],
                'biz_source': '广告投流',
                'exclude_data_sources': ['额外补充'],
            },
            'manpower': {
                'monthly_business_unit': '拍摄',
                'weekly_business_unit': '拍摄',
                'department_scope': ['品牌设计部', '内容二部'],
                'department_position_map': {'品牌设计部': '剪辑', '内容二部': '剪辑'},
            },
            'metrics': {
                'output_field': '产出数',
                'enable_project_count': False,
                'exclude_intern_efficiency': False,
                'show_total': False,
                'hidden': False,
            },
        },
    ]

    # 这一组 helper 的作用是把“配置长什么样”和“业务流程怎么读配置”解耦。
    # 后面如果 MODULES 再调整结构，优先改这些函数，而不是满文件搜字段名。
    def get_module_filter_options(module_config):
        return module_config['filters']

    def get_module_manpower_options(module_config):
        return module_config['manpower']

    def get_module_metric_options(module_config):
        return module_config['metrics']

    def get_module_name(module_config):
        return module_config['name']

    def should_hide_module(module_config):
        return get_module_metric_options(module_config).get('hidden', False)

    def should_show_total(module_config):
        return get_module_metric_options(module_config).get('show_total', True)

    def get_module_output_field(module_config):
        return get_module_metric_options(module_config).get('output_field', '产出数')

    def get_module_alternate_output_field(module_config):
        return get_module_metric_options(module_config).get('alternate_output_field')

    def get_module_efficiency_output_field(module_config):
        return get_module_metric_options(module_config).get('efficiency_output_field', get_module_output_field(module_config))

    def should_enable_project_count(module_config):
        return get_module_metric_options(module_config).get('enable_project_count', False)

    def should_exclude_intern_efficiency(module_config):
        return get_module_metric_options(module_config).get('exclude_intern_efficiency', False)

    def get_module_business_units(module_config):
        return get_module_filter_options(module_config)['business_units']

    def get_module_posts(module_config):
        return get_module_filter_options(module_config)['posts']

    def get_module_business_source(module_config):
        return get_module_filter_options(module_config)['biz_source']

    def get_module_excluded_data_sources(module_config):
        return get_module_filter_options(module_config).get('exclude_data_sources')

    def get_department_scope(module_config):
        return get_module_manpower_options(module_config).get('department_scope')

    def get_department_position_map(module_config):
        return get_module_manpower_options(module_config)['department_position_map']

    def get_monthly_manpower_business_unit(module_config):
        return get_module_manpower_options(module_config)['monthly_business_unit']

    def get_weekly_manpower_business_unit(module_config):
        return get_module_manpower_options(module_config)['weekly_business_unit']

    def build_module_runtime_options(module_config):
        """把主流程反复要用的模块配置整理成一份已解释好的运行参数。"""
        return {
            'name': get_module_name(module_config),
            'show_total': should_show_total(module_config),
            'department_scope': get_department_scope(module_config),
            'department_position_map': get_department_position_map(module_config),
            'monthly_manpower_business_unit': get_monthly_manpower_business_unit(module_config),
            'weekly_manpower_business_unit': get_weekly_manpower_business_unit(module_config),
            'output_field': get_module_output_field(module_config),
            'alternate_output_field': get_module_alternate_output_field(module_config),
            'efficiency_output_field': get_module_efficiency_output_field(module_config),
            'enable_project_count': should_enable_project_count(module_config),
            'exclude_intern_efficiency': should_exclude_intern_efficiency(module_config),
            'business_units': get_module_business_units(module_config),
            'posts': get_module_posts(module_config),
            'business_source': get_module_business_source(module_config),
            'exclude_data_sources': get_module_excluded_data_sources(module_config),
        }

    def get_sorted_departments(module_options):
        """部门排序规则保持不变，只是单独收成函数，方便后面理解和替换。"""
        department_scope = module_options['department_scope']
        department_position_map = module_options['department_position_map']

        if department_scope is not None:
            return department_scope

        dept_output = {}
        for dept in department_position_map:
            total = 0
            for month_key in RECENT_MONTHS:
                rows = filter_source_rows(
                    business_units=module_options['business_units'],
                    posts=module_options['posts'],
                    biz_source=module_options['business_source'],
                    exclude_data_sources=module_options['exclude_data_sources'],
                    months=[month_key],
                    depts=[dept]
                )
                total += sum_output(rows, module_options['output_field'])
            dept_output[dept] = total

        return sorted(dept_output.keys(), key=lambda d: dept_output[d], reverse=True)

    def apply_efficiency_adjustments(metrics, rows, workday_count, workload_manpower, module_options):
        """效率指标是否排除实习生，完全由模块配置控制。"""
        if not module_options['exclude_intern_efficiency'] or not rows:
            return metrics

        non_intern_rows = exclude_intern_rows(rows)
        if not non_intern_rows:
            return metrics

        efficiency_metrics = calculate_efficiency_metrics(
            non_intern_rows,
            workday_count,
            workload_manpower,
            module_options['efficiency_output_field']
        )
        metrics['avg_daily_workload'] = efficiency_metrics['avg_daily_workload']
        metrics['saturation'] = efficiency_metrics['saturation']
        metrics['single_time'] = efficiency_metrics['single_time']
        return metrics

    def attach_project_metrics(metrics, rows, output_manpower, module_options):
        """项目数是显式配置，不再通过 output_field 之类的隐式规则推断。"""
        if not module_options['enable_project_count']:
            return metrics

        project_count = count_distinct_projects(rows)
        metrics['project_count'] = project_count
        metrics['project_per_capita'] = round(project_count / output_manpower, 1) if output_manpower > 0 else 0
        return metrics

    def attach_alternate_output_metrics(metrics, rows, workday_count, output_manpower, module_options):
        """编剧产出区口径切换：对接脚本数直接使用产出数列。"""
        alternate_output_field = module_options.get('alternate_output_field')
        if not alternate_output_field:
            return metrics

        alternate_total_output = sum_output(rows, alternate_output_field)
        if workday_count == 0 or output_manpower == 0:
            alternate_avg_daily_output = 0
        else:
            alternate_avg_daily_output = alternate_total_output / workday_count / output_manpower

        metrics['handoff_total_output'] = round(alternate_total_output, 2)
        metrics['handoff_avg_daily_output'] = round(alternate_avg_daily_output, 4)
        return metrics

    def attach_total_project_metrics(metrics, detail_metrics, period_key, sorted_depts, total_output_manpower, module_options):
        """总计项目数来自各部门项目数汇总，避免月度/周度重复写同一段逻辑。"""
        if not module_options['enable_project_count']:
            return metrics

        total_project_count = sum(detail_metrics[dept][period_key].get('project_count', 0) for dept in sorted_depts)
        metrics['project_count'] = total_project_count
        metrics['project_per_capita'] = round(total_project_count / total_output_manpower, 1) if total_output_manpower > 0 else 0
        return metrics

    def calculate_monthly_metrics(module_options, module_name, sorted_depts, month_periods=None):
        month_periods = month_periods or RECENT_MONTHS
        department_position_map = module_options['department_position_map']
        monthly_manpower_business_unit = module_options['monthly_manpower_business_unit']

        dept_month_metrics = {}
        for dept in sorted_depts:
            post = department_position_map[dept]
            dept_month_metrics[dept] = {}
            for month_key in month_periods:
                rows = filter_source_rows(
                    business_units=module_options['business_units'],
                    posts=module_options['posts'],
                    biz_source=module_options['business_source'],
                    exclude_data_sources=module_options['exclude_data_sources'],
                    months=[month_key],
                    depts=[dept]
                )
                output_manpower = get_manpower_value(monthly_manpower_rows, month_key, dept, post,
                                                     monthly_manpower_business_unit, '产出用')
                workload_manpower = get_manpower_value(monthly_manpower_rows, month_key, dept, post,
                                                       monthly_manpower_business_unit, '工作量用')
                workday_count = month_workdays.get(month_key, 22)
                debug_info = {'type': '月度', 'module': module_name, 'dept': dept, 'period': month_key}
                metrics = calculate_metrics(rows, workday_count, output_manpower, workload_manpower,
                                            module_options['output_field'], debug_info,
                                            module_options['efficiency_output_field'])
                metrics = apply_efficiency_adjustments(metrics, rows, workday_count, workload_manpower, module_options)
                metrics = attach_project_metrics(metrics, rows, output_manpower, module_options)
                metrics = attach_alternate_output_metrics(metrics, rows, workday_count, output_manpower, module_options)
                dept_month_metrics[dept][month_key] = metrics

        total_month_metrics = {}
        valid_depts = list(department_position_map.keys())
        for month_key in month_periods:
            rows = filter_source_rows(
                business_units=module_options['business_units'],
                posts=module_options['posts'],
                biz_source=module_options['business_source'],
                exclude_data_sources=module_options['exclude_data_sources'],
                months=[month_key],
                depts=valid_depts
            )
            total_output_manpower = sum(
                get_manpower_value(monthly_manpower_rows, month_key, dept, department_position_map[dept],
                                   monthly_manpower_business_unit, '产出用')
                for dept in valid_depts
            )
            total_workload_manpower = sum(
                get_manpower_value(monthly_manpower_rows, month_key, dept, department_position_map[dept],
                                   monthly_manpower_business_unit, '工作量用')
                for dept in valid_depts
            )
            workday_count = month_workdays.get(month_key, 22)
            debug_info = {'type': '月度-总计', 'module': module_name, 'dept': '总计', 'period': month_key}
            metrics = calculate_metrics(rows, workday_count, total_output_manpower, total_workload_manpower,
                                        module_options['output_field'], debug_info,
                                        module_options['efficiency_output_field'])
            metrics = apply_efficiency_adjustments(metrics, rows, workday_count, total_workload_manpower, module_options)
            metrics = attach_total_project_metrics(
                metrics,
                dept_month_metrics,
                month_key,
                sorted_depts,
                total_output_manpower,
                module_options
            )
            metrics = attach_alternate_output_metrics(metrics, rows, workday_count, total_output_manpower, module_options)
            total_month_metrics[month_key] = metrics

        if module_name in LABOR_COST_OFFSET:
            offsets = LABOR_COST_OFFSET[module_name]
            for dept in sorted_depts:
                offset = offsets.get(dept, 0)
                post = department_position_map[dept]
                for month_key in month_periods:
                    output_manpower = get_manpower_value(monthly_manpower_rows, month_key, dept, post,
                                                         monthly_manpower_business_unit, '产出用')
                    cost_per_capita = get_labor_cost_per_capita(month_key, dept)
                    total_output = dept_month_metrics[dept][month_key]['total_output']
                    labor_cost_numerator = (output_manpower + offset) * cost_per_capita
                    if total_output > 0:
                        dept_month_metrics[dept][month_key]['single_labor_cost'] = round(
                            labor_cost_numerator / total_output, 2)
                    else:
                        dept_month_metrics[dept][month_key]['single_labor_cost'] = 0
                    dept_month_metrics[dept][month_key]['labor_cost_numerator'] = round(labor_cost_numerator, 2)
                    dept_month_metrics[dept][month_key]['labor_cost_per_capita'] = round(cost_per_capita, 2)
                    dept_month_metrics[dept][month_key]['labor_cost_offset'] = round(offset, 2)

            for month_key in month_periods:
                total_output = total_month_metrics[month_key]['total_output']
                manpower_items = []
                total_labor_cost_numerator = 0
                for dept in valid_depts:
                    offset = offsets.get(dept, 0)
                    post = department_position_map[dept]
                    output_manpower = get_manpower_value(monthly_manpower_rows, month_key, dept, post,
                                                         monthly_manpower_business_unit, '产出用')
                    labor_cost_per_capita = get_labor_cost_per_capita(month_key, dept)
                    manpower_items.append((output_manpower, offset, labor_cost_per_capita))
                    total_labor_cost_numerator += (output_manpower + offset) * labor_cost_per_capita
                total_month_metrics[month_key]['single_labor_cost'] = round(
                    calc_single_labor_cost(total_output, manpower_items), 2)
                total_month_metrics[month_key]['labor_cost_numerator'] = round(total_labor_cost_numerator, 2)

        return dept_month_metrics, total_month_metrics

    def calculate_weekly_metrics(module_options, module_name, sorted_depts):
        department_position_map = module_options['department_position_map']
        weekly_manpower_business_unit = module_options['weekly_manpower_business_unit']
        week_key_getter = get_content_editing_week_key if module_name == '内容团队-剪辑' else None

        dept_week_metrics = {}
        for dept in sorted_depts:
            post = department_position_map[dept]
            dept_week_metrics[dept] = {}
            for week_key in RECENT_WEEKS:
                rows = filter_source_rows(
                    business_units=module_options['business_units'],
                    posts=module_options['posts'],
                    biz_source=module_options['business_source'],
                    exclude_data_sources=module_options['exclude_data_sources'],
                    weeks=[week_key],
                    depts=[dept],
                    week_key_getter=week_key_getter
                )
                output_manpower = get_manpower_value(weekly_manpower_rows, week_key, dept, post,
                                                     weekly_manpower_business_unit, '产出用')
                workload_manpower = get_manpower_value(weekly_manpower_rows, week_key, dept, post,
                                                       weekly_manpower_business_unit, '工作量用')
                workday_count = week_workdays.get(week_key, 5)
                debug_info = {'type': '周度', 'module': module_name, 'dept': dept, 'period': week_key}
                metrics = calculate_metrics(rows, workday_count, output_manpower, workload_manpower,
                                            module_options['output_field'], debug_info,
                                            module_options['efficiency_output_field'])
                metrics = apply_efficiency_adjustments(metrics, rows, workday_count, workload_manpower, module_options)
                metrics = attach_project_metrics(metrics, rows, output_manpower, module_options)
                metrics = attach_alternate_output_metrics(metrics, rows, workday_count, output_manpower, module_options)
                dept_week_metrics[dept][week_key] = metrics

        total_week_metrics = {}
        valid_depts = list(department_position_map.keys())
        for week_key in RECENT_WEEKS:
            rows = filter_source_rows(
                business_units=module_options['business_units'],
                posts=module_options['posts'],
                biz_source=module_options['business_source'],
                exclude_data_sources=module_options['exclude_data_sources'],
                weeks=[week_key],
                depts=valid_depts,
                week_key_getter=week_key_getter
            )
            total_output_manpower = sum(
                get_manpower_value(weekly_manpower_rows, week_key, dept, department_position_map[dept],
                                   weekly_manpower_business_unit, '产出用')
                for dept in valid_depts
            )
            total_workload_manpower = sum(
                get_manpower_value(weekly_manpower_rows, week_key, dept, department_position_map[dept],
                                   weekly_manpower_business_unit, '工作量用')
                for dept in valid_depts
            )
            workday_count = week_workdays.get(week_key, 5)
            debug_info = {'type': '周度-总计', 'module': module_name, 'dept': '总计', 'period': week_key}
            metrics = calculate_metrics(rows, workday_count, total_output_manpower, total_workload_manpower,
                                        module_options['output_field'], debug_info,
                                        module_options['efficiency_output_field'])
            metrics = apply_efficiency_adjustments(metrics, rows, workday_count, total_workload_manpower, module_options)
            metrics = attach_total_project_metrics(
                metrics,
                dept_week_metrics,
                week_key,
                sorted_depts,
                total_output_manpower,
                module_options
            )
            metrics = attach_alternate_output_metrics(metrics, rows, workday_count, total_output_manpower, module_options)
            total_week_metrics[week_key] = metrics

        return dept_week_metrics, total_week_metrics

    def build_structure_view_data(module_options, periods, period_key, output_field=None):
        """结构分析复用同一套筛选规则，只是关注占比和波动而不是效率指标。"""
        structure_output_field = output_field or module_options['output_field']
        last_period = periods[-1]
        all_depts = ['全部'] + list(module_options['department_position_map'].keys())
        view_result = {}

        for dept_key in all_depts:
            rows_latest = filter_source_rows(
                business_units=module_options['business_units'],
                posts=module_options['posts'],
                biz_source=module_options['business_source'],
                exclude_data_sources=module_options['exclude_data_sources'],
                depts=None if dept_key == '全部' else [dept_key],
                **{period_key: [last_period]}
            )

            center_map = {}
            for row in rows_latest:
                center = (row.get('修正所属中心') or '未知').strip()
                dept_req = (row.get('修正需求部门') or '未知').strip()
                out = row.get(structure_output_field, 0) or 0

                if center in EXPAND_CENTERS:
                    item_key = center + '|' + dept_req
                    if item_key not in center_map:
                        center_map[item_key] = {'label': dept_req, 'group': center, 'output': 0, 'expanded': True}
                    center_map[item_key]['output'] += out
                else:
                    if center not in center_map:
                        center_map[center] = {'label': center, 'group': center, 'output': 0, 'expanded': False}
                    center_map[center]['output'] += out

            share_list = sorted(center_map.values(), key=lambda x: x['output'], reverse=True)
            total_output_share = sum(item['output'] for item in share_list)

            if len(periods) >= 2:
                prev_period = periods[-2]
                rows_prev = filter_source_rows(
                    business_units=module_options['business_units'],
                    posts=module_options['posts'],
                    biz_source=module_options['business_source'],
                    exclude_data_sources=module_options['exclude_data_sources'],
                    depts=None if dept_key == '全部' else [dept_key],
                    **{period_key: [prev_period]}
                )

                def build_project_map(rows):
                    project_map = {}
                    for row in rows:
                        center = (row.get('修正所属中心') or '').strip()
                        dept_req = (row.get('修正需求部门') or '').strip()
                        product = (row.get('修正产品') or '').strip()
                        out = row.get(structure_output_field, 0) or 0
                        project_key = f"{center} / {dept_req} / {product}" if center and dept_req and product else '未知项目'
                        project_map[project_key] = project_map.get(project_key, 0) + out
                    return project_map

                prev_map = build_project_map(rows_prev)
                curr_map = build_project_map(rows_latest)

                all_keys = set(list(prev_map.keys()) + list(curr_map.keys()))
                changes = []
                for project_key in all_keys:
                    prev_value = prev_map.get(project_key, 0)
                    curr_value = curr_map.get(project_key, 0)
                    diff = curr_value - prev_value
                    pct = round(diff / prev_value * 100, 1) if prev_value > 0 else None
                    changes.append({
                        'project': project_key,
                        'prev': prev_value,
                        'curr': curr_value,
                        'diff': diff,
                        'pct': pct,
                    })

                increases = [change for change in changes if change['diff'] > 0]
                decreases = [change for change in changes if change['diff'] < 0]
                increases.sort(key=lambda x: x['diff'], reverse=True)
                decreases.sort(key=lambda x: x['diff'])
                top_up = increases[:3]
                top_down = decreases[:3]
            else:
                top_up = []
                top_down = []

            view_result[dept_key] = {
                'share_list': share_list,
                'total_output_share': total_output_share,
                'top_up': top_up,
                'top_down': top_down,
            }

        return view_result


    def build_efficiency_view_data(module_options, periods, period_key):
        output_field = module_options['efficiency_output_field']
        exclude_intern = module_options['exclude_intern_efficiency']
        sorted_depts = get_sorted_departments(module_options)
        latest_period = periods[-1] if periods else None
        prev_period = periods[-2] if len(periods) >= 2 else None
        workday_map = month_workdays if period_key == 'months' else week_workdays
        view_result = {}

        for dept_key in ['全部'] + sorted_depts:
            rows_latest = filter_source_rows(
                business_units=module_options['business_units'],
                posts=module_options['posts'],
                biz_source=module_options['business_source'],
                exclude_data_sources=module_options['exclude_data_sources'],
                depts=None if dept_key == '全部' else [dept_key],
                **{period_key: [latest_period]} if latest_period else {}
            )
            rows_latest = exclude_intern_rows(rows_latest) if exclude_intern else rows_latest

            def build_demand_dept_time(rows):
                demand_map = {}
                for row in rows:
                    demand_dept = (row.get('修正需求部门') or '未知').strip() or '未知'
                    output_value = row.get(output_field, 0) or 0
                    workload_value = row.get('周报工作量', 0) or 0
                    if demand_dept not in demand_map:
                        demand_map[demand_dept] = {'label': demand_dept, 'output': 0, 'workload': 0}
                    demand_map[demand_dept]['output'] += output_value
                    demand_map[demand_dept]['workload'] += workload_value

                result = []
                for item in demand_map.values():
                    single_time = item['workload'] * 60 / item['output'] if item['output'] > 0 else 0
                    result.append({
                        'label': item['label'],
                        'single_time': round(single_time, 2),
                        'output': round(item['output'], 2),
                    })
                result.sort(key=lambda x: x['single_time'], reverse=True)
                return result

            def build_project_time_map(rows):
                project_map = {}
                for row in rows:
                    center = (row.get('修正所属中心') or '').strip()
                    demand_dept = (row.get('修正需求部门') or '').strip()
                    product = (row.get('修正产品') or '').strip()
                    project_key = f"{center} / {demand_dept} / {product}" if center and demand_dept and product else '未知项目'
                    output_value = row.get(output_field, 0) or 0
                    workload_value = row.get('周报工作量', 0) or 0
                    if project_key not in project_map:
                        project_map[project_key] = {'project': project_key, 'output': 0, 'workload': 0}
                    project_map[project_key]['output'] += output_value
                    project_map[project_key]['workload'] += workload_value

                result = {}
                for project_key, item in project_map.items():
                    if item['output'] <= 0:
                        continue
                    result[project_key] = {
                        'project': project_key,
                        'output': round(item['output'], 2),
                        'single_time': round(item['workload'] * 60 / item['output'], 2),
                    }
                return result

            total_latest_output = sum(row.get(output_field, 0) or 0 for row in rows_latest)
            total_latest_workload = sum(row.get('周报工作量', 0) or 0 for row in rows_latest)
            overall_curr_single_time = round(total_latest_workload * 60 / total_latest_output, 2) if total_latest_output > 0 else 0
            demand_dept_times = build_demand_dept_time(rows_latest)
            for item in demand_dept_times:
                output_share = item['output'] / total_latest_output * 100 if total_latest_output > 0 else 0
                item['output_share'] = round(output_share, 1)
            latest_project_map = build_project_time_map(rows_latest)
            ranked_projects = sorted(latest_project_map.values(), key=lambda x: x['single_time'], reverse=True)
            top_high = ranked_projects[:5]
            top_low = sorted(ranked_projects, key=lambda x: x['single_time'])[:5]

            if prev_period:
                rows_prev = filter_source_rows(
                    business_units=module_options['business_units'],
                    posts=module_options['posts'],
                    biz_source=module_options['business_source'],
                    exclude_data_sources=module_options['exclude_data_sources'],
                    depts=None if dept_key == '全部' else [dept_key],
                    **{period_key: [prev_period]}
                )
                rows_prev = exclude_intern_rows(rows_prev) if exclude_intern else rows_prev
                total_prev_output = sum(row.get(output_field, 0) or 0 for row in rows_prev)
                total_prev_workload = sum(row.get('周报工作量', 0) or 0 for row in rows_prev)
                overall_prev_single_time = round(total_prev_workload * 60 / total_prev_output, 2) if total_prev_output > 0 else 0
                prev_project_map = build_project_time_map(rows_prev)

                prev_demand_dept_times = {item['label']: item for item in build_demand_dept_time(rows_prev)}
                curr_demand_dept_times = {item['label']: item for item in demand_dept_times}
                dept_impact_up = []
                dept_impact_down = []
                all_demand_depts = set(prev_demand_dept_times.keys()) | set(curr_demand_dept_times.keys())
                for demand_dept in all_demand_depts:
                    prev_info = prev_demand_dept_times.get(demand_dept, {'single_time': 0, 'output': 0})
                    curr_info = curr_demand_dept_times.get(demand_dept, {'single_time': 0, 'output': 0})
                    weighted_output = (prev_info['output'] + curr_info['output']) / 2
                    impact_minutes = (curr_info['single_time'] - prev_info['single_time']) * weighted_output
                    impact_item = {
                        'label': demand_dept,
                        'prev_single_time': round(prev_info['single_time'], 2),
                        'curr_single_time': round(curr_info['single_time'], 2),
                        'prev_output': round(prev_info['output'], 2),
                        'curr_output': round(curr_info['output'], 2),
                        'weighted_output': round(weighted_output, 2),
                        'impact_minutes': round(impact_minutes, 1),
                    }
                    if impact_minutes > 0:
                        dept_impact_up.append(impact_item)
                    elif impact_minutes < 0:
                        dept_impact_down.append(impact_item)
                dept_impact_up.sort(key=lambda x: x['impact_minutes'], reverse=True)
                dept_impact_down.sort(key=lambda x: x['impact_minutes'])
                top_dept_impact_up = dept_impact_up[:3]
                top_dept_impact_down = dept_impact_down[:3]

                all_project_keys = set(prev_project_map.keys()) | set(latest_project_map.keys())
                top_fluctuation = []
                project_impact_up = []
                project_impact_down = []
                for project_key in all_project_keys:
                    prev_info = prev_project_map.get(project_key)
                    latest_info = latest_project_map.get(project_key)
                    prev_single_time = prev_info['single_time'] if prev_info else 0
                    latest_single_time = latest_info['single_time'] if latest_info else 0
                    diff_value = latest_single_time - prev_single_time
                    top_fluctuation.append({
                        'project': project_key,
                        'prev': round(prev_single_time, 2),
                        'curr': round(latest_single_time, 2),
                        'diff': round(diff_value, 2),
                        'abs_diff': round(abs(diff_value), 2),
                    })
                    weighted_output = ((prev_info['output'] if prev_info else 0) + (latest_info['output'] if latest_info else 0)) / 2
                    impact_minutes = diff_value * weighted_output
                    impact_item = {
                        'project': project_key,
                        'prev_single_time': round(prev_single_time, 2),
                        'curr_single_time': round(latest_single_time, 2),
                        'prev_output': round(prev_info['output'], 2) if prev_info else 0,
                        'curr_output': round(latest_info['output'], 2) if latest_info else 0,
                        'weighted_output': round(weighted_output, 2),
                        'impact_minutes': round(impact_minutes, 1),
                    }
                    if impact_minutes > 0:
                        project_impact_up.append(impact_item)
                    elif impact_minutes < 0:
                        project_impact_down.append(impact_item)
                top_fluctuation.sort(key=lambda x: x['abs_diff'], reverse=True)
                top_fluctuation = top_fluctuation[:5]
                project_impact_up.sort(key=lambda x: x['impact_minutes'], reverse=True)
                project_impact_down.sort(key=lambda x: x['impact_minutes'])
                top_project_impact_up = project_impact_up[:3]
                top_project_impact_down = project_impact_down[:3]
            else:
                overall_prev_single_time = 0
                top_dept_impact_up = []
                top_dept_impact_down = []
                top_project_impact_up = []
                top_project_impact_down = []
                top_fluctuation = []

            view_result[dept_key] = {
                'latest_period': latest_period,
                'workdays': workday_map.get(latest_period, 0) if latest_period else 0,
                'overall_prev_single_time': overall_prev_single_time,
                'overall_curr_single_time': overall_curr_single_time,
                'demand_dept_times': demand_dept_times,
                'top_dept_impact_up': top_dept_impact_up,
                'top_dept_impact_down': top_dept_impact_down,
                'top_project_impact_up': top_project_impact_up,
                'top_project_impact_down': top_project_impact_down,
                'top_high': top_high,
                'top_low': top_low,
                'top_fluctuation': top_fluctuation,
            }

        return view_result

    # ============================================================
    # 异常简析数据
    # ============================================================
    def safe_pct(curr, prev):
        if prev in (0, None):
            return None
        return (curr - prev) / prev * 100

    def direction_word(diff):
        return '增加' if diff > 0 else '减少'

    def format_qty(value, unit):
        value = abs(value)
        if value >= 100:
            text = f'{value:.0f}'
        elif value >= 10:
            text = f'{value:.1f}'.rstrip('0').rstrip('.')
        else:
            text = f'{value:.1f}'.rstrip('0').rstrip('.')
        return f'{text}{unit}'

    def format_people(value):
        return f'{abs(value):.1f}'.rstrip('0').rstrip('.') + '人'

    def format_money(value):
        value = abs(value)
        if value >= 10000:
            return f'{value / 10000:.1f}'.rstrip('0').rstrip('.') + '万'
        return f'{value:.0f}'

    def anomaly_project_html(label):
        return f'<strong class="anomaly-project">{html_lib.escape(str(label), quote=False)}</strong>'

    def anomaly_delta_html(value, text):
        tone = 'up' if value > 0 else 'down'
        return f'<span class="anomaly-delta {tone}">{html_lib.escape(str(text), quote=False)}</span>'

    def anomaly_metric_headline(view, diff_pct, up_word='增加', down_word='减少'):
        if diff_pct is None:
            return ''
        period_word = '周环比' if view == 'weekly' else '月环比'
        trend = up_word if diff_pct > 0 else down_word
        return (
            f'{period_word}{trend}' +
            anomaly_delta_html(diff_pct, f'{abs(diff_pct):.1f}%')
        )

    def get_output_unit(module_name):
        if module_name == '图片':
            return '张'
        if module_name == '内容团队-编剧':
            return '个'
        return '条'

    def get_project_label(row):
        demand_dept = (row.get('修正需求部门') or '').strip()
        product = (row.get('修正产品') or row.get('修正产品名称') or '').strip()
        if demand_dept and product:
            return f'{demand_dept} · {product}'
        if demand_dept:
            return demand_dept
        if product:
            return product
        return '未知项目'

    def get_project_key(row):
        demand_dept = (row.get('修正需求部门') or '').strip()
        product = (row.get('修正产品') or row.get('修正产品名称') or '').strip()
        return (demand_dept, product)

    def get_rows_for_period(module_options, view, period, dept):
        period_key = 'months' if view == 'monthly' else 'weeks'
        week_key_getter = get_content_editing_week_key if module_options['name'] == '内容团队-剪辑' else None
        return filter_source_rows(
            business_units=module_options['business_units'],
            posts=module_options['posts'],
            biz_source=module_options['business_source'],
            exclude_data_sources=module_options['exclude_data_sources'],
            depts=[dept],
            week_key_getter=week_key_getter,
            **{period_key: [period]}
        )

    def build_project_output_map(rows, output_field):
        project_map = {}
        for row in rows:
            key = get_project_key(row)
            if key not in project_map:
                project_map[key] = {'label': get_project_label(row), 'output': 0}
            project_map[key]['output'] += row.get(output_field, 0) or 0
        return project_map

    def build_project_efficiency_map(rows, output_field):
        project_map = {}
        for row in rows:
            key = get_project_key(row)
            if key not in project_map:
                project_map[key] = {'label': get_project_label(row), 'output': 0, 'workload_minutes': 0}
            project_map[key]['output'] += row.get(output_field, 0) or 0
            project_map[key]['workload_minutes'] += (row.get('周报工作量', 0) or 0) * 60
        for item in project_map.values():
            item['single_time'] = item['workload_minutes'] / item['output'] if item['output'] > 0 else 0
        return project_map

    SIMPLE_DEMAND_TYPES = {
        '图片': ['轮播小图', '简单修改', '设计找图', '宫格小图'],
        '混剪': ['混剪修改'],
    }

    def is_simple_demand_type(module_name, demand_type):
        simple_types = SIMPLE_DEMAND_TYPES.get(module_name, [])
        demand_type_text = str(demand_type or '').strip()
        return any(simple_type in demand_type_text for simple_type in simple_types)

    def build_period_meta(view, periods):
        if view == 'weekly':
            if len(periods) < 2:
                return {'analysis_period': None, 'compare_period': None, 'label': '周度样本不足'}
            curr, prev = periods[-1], periods[-2]
            return {
                'analysis_period': curr,
                'compare_period': prev,
                'skipped_period': None,
                'label': f'{curr} 对比 {prev}',
            }

        if len(periods) < 2:
            return {'analysis_period': None, 'compare_period': None, 'label': '月度样本不足'}

        latest_month = periods[-1]
        _, latest_month_end = get_month_range(latest_month)
        latest_complete = bool(latest_week_end_date and latest_week_end_date >= latest_month_end)
        if latest_complete or len(periods) < 3:
            curr, prev = periods[-1], periods[-2]
            return {
                'analysis_period': curr,
                'compare_period': prev,
                'skipped_period': None,
                'label': f'{curr} 对比 {prev}',
            }

        curr, prev = periods[-2], periods[-3]
        return {
            'analysis_period': curr,
            'compare_period': prev,
            'skipped_period': latest_month,
            'skipped_reason': 'latest_month_incomplete',
            'label': f'{curr} 对比 {prev}（{latest_month} 未完结，暂不参与）',
        }

    def build_total_output_reason(module_options, view, dept, curr_period, prev_period, output_field, unit):
        curr_rows = get_rows_for_period(module_options, view, curr_period, dept)
        prev_rows = get_rows_for_period(module_options, view, prev_period, dept)
        curr_total = sum_output(curr_rows, output_field)
        prev_total = sum_output(prev_rows, output_field)
        total_diff = curr_total - prev_total
        total_pct = safe_pct(curr_total, prev_total)
        if total_pct is None or abs(total_pct) < 5:
            return ''

        curr_map = build_project_output_map(curr_rows, output_field)
        prev_map = build_project_output_map(prev_rows, output_field)
        items = []
        for key in set(curr_map.keys()) | set(prev_map.keys()):
            curr = curr_map.get(key, {'output': 0, 'label': prev_map.get(key, {}).get('label', '未知项目')})
            prev = prev_map.get(key, {'output': 0, 'label': curr.get('label', '未知项目')})
            diff = curr['output'] - prev['output']
            if total_diff > 0 and diff <= 0:
                continue
            if total_diff < 0 and diff >= 0:
                continue
            items.append({
                'label': curr.get('label') or prev.get('label') or '未知项目',
                'diff': diff,
                'share': abs(diff) / abs(total_diff) if total_diff else 0,
            })

        items.sort(key=lambda x: abs(x['diff']), reverse=True)
        if not items:
            return ''

        def item_text(item):
            return (
                anomaly_project_html(item['label']) +
                direction_word(item['diff']) +
                anomaly_delta_html(item['diff'], format_qty(item['diff'], unit))
            )

        first = items[0]
        second = items[1] if len(items) > 1 else None
        third = items[2] if len(items) > 2 else None
        if first['share'] >= 0.35:
            project_text = item_text(first) + '。'
        elif first['share'] >= 0.15:
            if second and second['share'] >= 0.10:
                project_text = item_text(first) + '、' + item_text(second) + '。'
            else:
                extras = [item for item in [second, third] if item]
                if extras:
                    project_text = item_text(first) + '，此外' + '、'.join(item_text(item) for item in extras) + '小量波动引起。'
                else:
                    project_text = item_text(first) + '。'
        else:
            small_items = items[:3] if len(items) >= 3 else items[:2]
            project_text = '、'.join(item_text(item) for item in small_items) + '小量波动引起。'

        return {
            'headline': anomaly_metric_headline(view, total_pct),
            'detail': project_text,
        }

    def build_avg_daily_output_reason(metrics_by_period, curr_period, prev_period):
        curr = metrics_by_period.get(curr_period, {})
        prev = metrics_by_period.get(prev_period, {})
        output_pct = safe_pct(curr.get('total_output', 0), prev.get('total_output', 0))
        avg_pct = safe_pct(curr.get('avg_daily_output', 0), prev.get('avg_daily_output', 0))
        if output_pct is None or avg_pct is None:
            return {'headline': '', 'detail': ''}

        same_direction = (output_pct >= 0 and avg_pct >= 0) or (output_pct < 0 and avg_pct < 0)
        if same_direction and abs(avg_pct - output_pct) < 10:
            return ''

        manpower_diff = (curr.get('output_manpower', 0) or 0) - (prev.get('output_manpower', 0) or 0)
        workday_diff = (curr.get('workdays', 0) or 0) - (prev.get('workdays', 0) or 0)
        reasons = []
        if abs(manpower_diff) >= 0.1:
            if manpower_diff > 0:
                reasons.append('产出用人力增加' + anomaly_delta_html(manpower_diff, format_people(manpower_diff)))
            else:
                reasons.append('产出用人力减少' + anomaly_delta_html(manpower_diff, format_people(manpower_diff)))
        if abs(workday_diff) >= 1:
            if workday_diff > 0:
                reasons.append('工作日增加' + anomaly_delta_html(workday_diff, f'{abs(workday_diff):.0f}天'))
            else:
                reasons.append('工作日减少' + anomaly_delta_html(workday_diff, f'{abs(workday_diff):.0f}天'))
        if not reasons:
            return ''

        if not same_direction:
            trend_text = '总产出上涨，但人均日均产出下降' if output_pct > 0 else '总产出下降，但人均日均产出上升'
            return trend_text + '，主要因' + '、'.join(reasons) + '。'
        diff_word = '高于' if abs(avg_pct) > abs(output_pct) else '低于'
        return f'人均日均产出变动幅度{diff_word}总产出，主要因' + '、'.join(reasons) + '。'

    def build_single_time_reason(module_options, view, dept, curr_period, prev_period):
        output_field = module_options['efficiency_output_field']
        curr_rows = get_rows_for_period(module_options, view, curr_period, dept)
        prev_rows = get_rows_for_period(module_options, view, prev_period, dept)
        if module_options['exclude_intern_efficiency']:
            curr_rows = exclude_intern_rows(curr_rows)
            prev_rows = exclude_intern_rows(prev_rows)

        curr_output = sum_output(curr_rows, output_field)
        prev_output = sum_output(prev_rows, output_field)
        curr_minutes = sum((row.get('周报工作量', 0) or 0) * 60 for row in curr_rows)
        prev_minutes = sum((row.get('周报工作量', 0) or 0) * 60 for row in prev_rows)
        curr_avg = curr_minutes / curr_output if curr_output > 0 else 0
        prev_avg = prev_minutes / prev_output if prev_output > 0 else 0
        pct = safe_pct(curr_avg, prev_avg)
        if pct is None or abs(pct) <= 10:
            return ''

        def build_simple_type_summary(rows):
            total_output = 0
            simple_output = 0
            simple_type_map = {}
            simple_project_type_map = {}
            for row in rows:
                output_value = row.get(output_field, 0) or 0
                total_output += output_value
                demand_type = str(row.get('需求类型') or '').strip()
                if not is_simple_demand_type(module_options['name'], demand_type):
                    continue
                simple_output += output_value
                if demand_type not in simple_type_map:
                    simple_type_map[demand_type] = {'label': demand_type, 'output': 0}
                simple_type_map[demand_type]['output'] += output_value
                key = get_project_key(row)
                project_type_key = (key, demand_type)
                if project_type_key not in simple_project_type_map:
                    simple_project_type_map[project_type_key] = {
                        'label': get_project_label(row),
                        'demand_type': demand_type,
                        'output': 0,
                    }
                simple_project_type_map[project_type_key]['output'] += output_value
            share = simple_output / total_output if total_output > 0 else 0
            return {
                'total_output': total_output,
                'simple_output': simple_output,
                'share': share,
                'type_map': simple_type_map,
                'project_type_map': simple_project_type_map,
            }

        def build_simple_type_reason():
            curr_summary = build_simple_type_summary(curr_rows)
            prev_summary = build_simple_type_summary(prev_rows)
            if curr_summary['total_output'] <= 0 or prev_summary['total_output'] <= 0:
                return None
            share_diff = curr_summary['share'] - prev_summary['share']
            if (pct > 0 and share_diff >= 0) or (pct < 0 and share_diff <= 0):
                return None

            curr_type_map = curr_summary['type_map']
            prev_type_map = prev_summary['type_map']
            type_items = []
            for key in set(curr_type_map.keys()) | set(prev_type_map.keys()):
                curr_item = curr_type_map.get(key, {'label': prev_type_map.get(key, {}).get('label', '未知类别'), 'output': 0})
                prev_item = prev_type_map.get(key, {'label': curr_item.get('label', '未知类别'), 'output': 0})
                diff = curr_item['output'] - prev_item['output']
                if pct > 0 and diff >= 0:
                    continue
                if pct < 0 and diff <= 0:
                    continue
                type_items.append({
                    'label': curr_item.get('label') or prev_item.get('label') or '未知类别',
                    'diff': diff,
                })
            if not type_items:
                return None

            curr_map = curr_summary['project_type_map']
            prev_map = prev_summary['project_type_map']
            target_items = []
            for key in set(curr_map.keys()) | set(prev_map.keys()):
                curr_item = curr_map.get(key, {'label': prev_map.get(key, {}).get('label', '未知项目'), 'output': 0})
                prev_item = prev_map.get(key, {'label': curr_item.get('label', '未知项目'), 'output': 0})
                diff = curr_item['output'] - prev_item['output']
                if pct > 0 and diff >= 0:
                    continue
                if pct < 0 and diff <= 0:
                    continue
                target_items.append({
                    'label': curr_item.get('label') or prev_item.get('label') or '未知项目',
                    'demand_type': curr_item.get('demand_type') or prev_item.get('demand_type') or '简单需求',
                    'diff': diff,
                })
            if not target_items:
                return None

            type_items.sort(key=lambda item: abs(item['diff']), reverse=True)
            target_items.sort(key=lambda item: abs(item['diff']), reverse=True)
            top_item = target_items[0]
            share_point_diff = share_diff * 100
            share_word = '下降' if share_diff < 0 else '提升'
            type_word = '下降' if share_diff < 0 else '上升'
            type_text = '、'.join(item['label'] for item in type_items[:2])
            project_word = '减少' if top_item['diff'] < 0 else '增加'
            return {
                'headline': anomaly_metric_headline(view, pct, '上升', '下降'),
                'detail': (
                    f"主要因简单的需求占比{share_word}"
                    f"{abs(share_point_diff):.1f}个百分点"
                    f"（{type_text}类别{type_word}）<br>"
                    f"如{top_item['demand_type']}类别的{anomaly_project_html(top_item['label'])}{project_word}"
                    f"{anomaly_delta_html(top_item['diff'], format_qty(top_item['diff'], unit))}。"
                ),
            }

        curr_map = build_project_efficiency_map(curr_rows, output_field)
        prev_map = build_project_efficiency_map(prev_rows, output_field)
        direction = 1 if pct > 0 else -1
        unit = get_output_unit(module_options['name'])

        simple_type_reason = build_simple_type_reason()
        if simple_type_reason:
            return simple_type_reason

        reason_groups = {}

        def get_relevant_time_and_avg(curr, prev):
            if curr['output'] > 0:
                return curr['single_time'], curr_avg
            return prev['single_time'], prev_avg

        def get_time_level(single_time, avg_time):
            if avg_time <= 0:
                return 'regular'
            if single_time >= avg_time * 1.5:
                return 'high'
            if single_time <= avg_time * 0.5:
                return 'low'
            return 'regular'

        def get_regular_relation(single_time, avg_time):
            return '高于平均水平' if single_time >= avg_time else '低于平均水平'

        def add_reason(group_key, group_label, item):
            if group_key not in reason_groups:
                reason_groups[group_key] = {'label': group_label, 'items': [], 'impact': 0}
            reason_groups[group_key]['items'].append(item)
            reason_groups[group_key]['impact'] += abs(item['impact'])

        for key in set(curr_map.keys()) | set(prev_map.keys()):
            curr = curr_map.get(key, {'label': prev_map.get(key, {}).get('label', '未知项目'), 'output': 0, 'workload_minutes': 0, 'single_time': 0})
            prev = prev_map.get(key, {'label': curr.get('label', '未知项目'), 'output': 0, 'workload_minutes': 0, 'single_time': 0})
            impact = curr['workload_minutes'] - prev['workload_minutes'] - prev_avg * (curr['output'] - prev['output'])
            if direction > 0 and impact <= 0:
                continue
            if direction < 0 and impact >= 0:
                continue

            output_delta = curr['output'] - prev['output']
            relevant_time, relevant_avg = get_relevant_time_and_avg(curr, prev)
            level = get_time_level(relevant_time, relevant_avg)
            relation = get_regular_relation(relevant_time, relevant_avg)
            item = {
                'label': curr.get('label') or prev.get('label') or '未知项目',
                'curr_output': curr['output'],
                'prev_output': prev['output'],
                'curr_single_time': curr['single_time'],
                'prev_single_time': prev['single_time'],
                'relevant_relation': relation,
                'output_delta': output_delta,
                'time_delta': curr['single_time'] - prev['single_time'],
                'impact': impact,
            }

            if direction > 0:
                if output_delta > 0 and level == 'high':
                    add_reason('high_increase', '高耗时项目产出增加', item)
                elif level == 'regular' and relevant_time >= relevant_avg:
                    add_reason('regular_high', '受到高于平均水平的常规项目影响', item)
                elif output_delta < 0 and level == 'low':
                    add_reason('low_decrease', '低耗时项目需求减少', item)
                elif level == 'regular' and relevant_time < relevant_avg:
                    add_reason('regular_low', '受到低于平均水平的常规项目影响', item)
            else:
                if output_delta > 0 and level == 'low':
                    add_reason('low_increase', '低耗时项目产出增加', item)
                elif level == 'regular' and relevant_time <= relevant_avg:
                    add_reason('regular_low', '受到低于平均水平的常规项目影响', item)
                elif output_delta < 0 and level == 'high':
                    add_reason('high_decrease', '高耗时项目需求减少', item)
                elif level == 'regular' and relevant_time > relevant_avg:
                    add_reason('regular_high', '受到高于平均水平的常规项目影响', item)

        groups = sorted(reason_groups.values(), key=lambda x: x['impact'], reverse=True)
        if not groups:
            return {'headline': '', 'detail': ''}
        selected_groups = groups[:1]

        def get_complex_note(item):
            output_delta = item['output_delta']
            time_delta = item['time_delta']
            if abs(time_delta) < 0.05 or output_delta == 0:
                return ''
            if direction > 0:
                if output_delta < 0 and time_delta > 0:
                    return '，虽产出减少，但单片耗时上升，总体推高部门平均耗时'
                if output_delta > 0 and time_delta < 0:
                    return '，虽单片耗时下降，但产出增加，总体推高部门平均耗时'
            else:
                if output_delta > 0 and time_delta > 0:
                    return '，虽单片耗时上升，但产出增加，总体拉低部门平均耗时'
                if output_delta < 0 and time_delta < 0:
                    return '，虽产出减少，但单片耗时下降，总体拉低部门平均耗时'
            return ''

        def project_text(item, show_complex_note=False):
            project = anomaly_project_html(item['label'])
            if item['curr_output'] == 0 and item['prev_output'] > 0:
                time_text = f"本期产出为0（上期为{item['prev_single_time']:.1f}分钟/{unit}）"
                output_text = f"产出量{format_qty(item['prev_output'], unit)} -> {format_qty(item['curr_output'], unit)}"
                return f"{project}{time_text}，{output_text}"
            if item['prev_output'] <= 0:
                time_text = f"{item['curr_single_time']:.1f}分钟/{unit}"
                note = get_complex_note(item) if show_complex_note else ''
                return f"{project}{time_text}，新项目产出{format_qty(item['curr_output'], unit)}{note}"
            if abs(item['curr_single_time'] - item['prev_single_time']) < 0.05:
                time_text = f"{item['curr_single_time']:.1f}分钟/{unit}"
            else:
                time_text = f"{item['prev_single_time']:.1f} -> {item['curr_single_time']:.1f}分钟/{unit}"
            output_text = f"产出量{format_qty(item['prev_output'], unit)} -> {format_qty(item['curr_output'], unit)}"
            note = get_complex_note(item) if show_complex_note else ''
            return f"{project}{time_text}，{output_text}{note}"

        def group_text(group):
            group['items'].sort(key=lambda x: abs(x['impact']), reverse=True)
            items = group['items'][:1]
            show_complex_note = '常规项目' in group['label']
            return (
                f"主要因{group['label']}<br>" +
                '；'.join(project_text(item, show_complex_note) for item in items) +
                '。'
            )

        return {
            'headline': anomaly_metric_headline(view, pct, '上升', '下降'),
            'detail': ''.join(group_text(group) for group in selected_groups),
        }

    def empty_anomaly_item():
        return {'headline': '', 'detail': '', 'text': ''}

    def normalize_anomaly_item(value):
        if isinstance(value, dict):
            headline = value.get('headline', '')
            detail = value.get('detail', '')
            text = value.get('text') or ((headline + '<br>' + detail) if headline and detail else headline or detail)
            return {'headline': headline, 'detail': detail, 'text': text}
        text = value or ''
        return {'headline': '', 'detail': text, 'text': text}

    def build_labor_cost_reason(module_options, dept_metrics, view, dept, curr_period, prev_period, output_field, unit):
        curr = dept_metrics.get(curr_period, {})
        prev = dept_metrics.get(prev_period, {})
        curr_cost = curr.get('single_labor_cost', 0) or 0
        prev_cost = prev.get('single_labor_cost', 0) or 0
        pct = safe_pct(curr_cost, prev_cost)
        if pct is None or abs(pct) < 10:
            return empty_anomaly_item()

        curr_num = curr.get('labor_cost_numerator', 0) or 0
        prev_num = prev.get('labor_cost_numerator', 0) or 0
        curr_output = curr.get('total_output', 0) or 0
        prev_output = prev.get('total_output', 0) or 0
        curr_labor_row = LABOR_COST_INDEX.get((curr_period, dept), {})
        prev_labor_row = LABOR_COST_INDEX.get((prev_period, dept), {})
        curr_labor_amount = curr_labor_row.get('人力+工位', 0) or 0
        prev_labor_amount = prev_labor_row.get('人力+工位', 0) or 0
        curr_total_manpower = curr_labor_row.get('总人力', 0) or 0
        prev_total_manpower = prev_labor_row.get('总人力', 0) or 0
        num_pct = safe_pct(curr_num, prev_num) or 0
        denom_pct = safe_pct(curr_output, prev_output) or 0
        labor_amount_pct = safe_pct(curr_labor_amount, prev_labor_amount)
        output_diff = curr_output - prev_output
        labor_amount_diff = curr_labor_amount - prev_labor_amount

        parts = []
        output_is_driver = abs(denom_pct) >= 8 and abs(denom_pct) >= abs(num_pct) * 0.75
        numerator_is_driver = abs(num_pct) >= 8 and abs(num_pct) >= abs(denom_pct) * 0.75

        if output_is_driver:
            project_reason = build_total_output_reason(module_options, view, dept, curr_period, prev_period, output_field, unit)
            project_reason = normalize_anomaly_item(project_reason).get('detail', '')
            project_reason = project_reason.rstrip('。')
            if project_reason:
                parts.append(
                    f'总产出{direction_word(output_diff)}' +
                    anomaly_delta_html(output_diff, format_qty(output_diff, unit)) +
                    f'，其中{project_reason}影响最大'
                )
            else:
                parts.append(
                    f'总产出{direction_word(output_diff)}' +
                    anomaly_delta_html(output_diff, format_qty(output_diff, unit))
                )

        labor_cost_can_describe = labor_amount_pct is not None and abs(labor_amount_pct) > 5
        if numerator_is_driver and labor_cost_can_describe:
            detail = []
            manpower_diff = curr_total_manpower - prev_total_manpower
            cost_pct = safe_pct(curr.get('labor_cost_per_capita', 0), prev.get('labor_cost_per_capita', 0))
            if abs(manpower_diff) >= 0.1:
                detail.append(
                    f"总人力{direction_word(manpower_diff)}" +
                    anomaly_delta_html(manpower_diff, format_people(manpower_diff))
                )
            if cost_pct is not None and abs(cost_pct) >= 1:
                detail.append(
                    f"人均人力成本{direction_word(cost_pct)}" +
                    anomaly_delta_html(cost_pct, f'{abs(cost_pct):.1f}%')
                )
            detail_text = '，其中' + '、'.join(detail) if detail else ''
            parts.append(
                f'人力成本{direction_word(labor_amount_diff)}' +
                anomaly_delta_html(labor_amount_diff, format_money(labor_amount_diff)) +
                detail_text
            )

        if not parts:
            return empty_anomaly_item()
        return {
            'headline': anomaly_metric_headline(view, pct, '上升', '下降'),
            'detail': '主要因' + '；'.join(parts) + '。',
        }

    def build_anomaly_reason_data(module_options, module_name, sorted_depts, module_data, labor_cost_module_data=None):
        result = {}
        unit = get_output_unit(module_name)
        for view in ['monthly', 'weekly']:
            periods = RECENT_MONTHS if view == 'monthly' else RECENT_WEEKS
            meta = build_period_meta(view, periods)
            curr_period = meta.get('analysis_period')
            prev_period = meta.get('compare_period')
            view_result = {'meta': meta, 'total_output': {}, 'avg_daily_output': {}, 'single_time': {}, 'single_labor_cost': {}}
            if not curr_period or not prev_period:
                result[view] = view_result
                continue

            for dept in sorted_depts:
                dept_metrics = module_data[view]['dept_metrics'][dept]
                view_result['total_output'][dept] = normalize_anomaly_item(
                    build_total_output_reason(
                        module_options, view, dept, curr_period, prev_period, module_options['output_field'], unit
                    )
                )
                view_result['avg_daily_output'][dept] = {
                    'text': build_avg_daily_output_reason(dept_metrics, curr_period, prev_period)
                }
                view_result['single_time'][dept] = normalize_anomaly_item(
                    build_single_time_reason(module_options, view, dept, curr_period, prev_period)
                )

                if view == 'monthly' and labor_cost_module_data:
                    cost_dept_metrics = labor_cost_module_data['dept_metrics'].get(dept, {})
                    if curr_period in cost_dept_metrics and prev_period in cost_dept_metrics:
                        view_result['single_labor_cost'][dept] = normalize_anomaly_item(
                            build_labor_cost_reason(
                                module_options, cost_dept_metrics, view, dept, curr_period, prev_period,
                                module_options['output_field'], unit
                            )
                        )
            result[view] = view_result
        if labor_cost_module_data and 'monthly' in result:
            monthly_cost = result['monthly'].get('single_labor_cost', {})
            monthly_meta = result['monthly'].get('meta', {})
            result.setdefault('weekly', {}).setdefault('single_labor_cost', {})
            result['weekly']['single_labor_cost'] = monthly_cost
            result['weekly']['labor_cost_meta'] = monthly_meta
        return result

    # ============================================================
    # 计算所有指标数据
    # ============================================================
    print('计算指标数据...')
    dashboard_data = {'monthly': {}, 'weekly': {}}
    labor_cost_data = {}

    for mod in MODULES:
        if should_hide_module(mod):
            continue
        module_options = build_module_runtime_options(mod)
        mname = module_options['name']
        print(f'  处理模块: {mname}')
        sorted_depts = get_sorted_departments(module_options)
        dept_month_metrics, total_month_metrics = calculate_monthly_metrics(module_options, mname, sorted_depts)
        dept_week_metrics, total_week_metrics = calculate_weekly_metrics(module_options, mname, sorted_depts)

        dashboard_data['monthly'][mname] = {
            'depts': sorted_depts,
            'dept_metrics': dept_month_metrics,
            'total_metrics': total_month_metrics,
            'show_total': module_options['show_total'],
        }
        dashboard_data['weekly'][mname] = {
            'depts': sorted_depts,
            'dept_metrics': dept_week_metrics,
            'total_metrics': total_week_metrics,
            'show_total': module_options['show_total'],
        }
        if mname in LABOR_COST_OFFSET:
            dept_cost_metrics, total_cost_metrics = calculate_monthly_metrics(
                module_options, mname, sorted_depts, LABOR_COST_MONTHS
            )
            labor_cost_data[mname] = {
                'depts': sorted_depts,
                'dept_metrics': dept_cost_metrics,
                'total_metrics': total_cost_metrics,
                'show_total': module_options['show_total'],
            }

    print('  数据计算完成')

    # ============================================================
    # 异常简析数据计算
    # ============================================================
    print('计算异常简析数据...')
    anomaly_reason_data = {'monthly': {}, 'weekly': {}}
    for mod in MODULES:
        if should_hide_module(mod):
            continue
        module_options = build_module_runtime_options(mod)
        mname = module_options['name']
        sorted_depts = get_sorted_departments(module_options)
        module_reason_data = build_anomaly_reason_data(
            module_options,
            mname,
            sorted_depts,
            {
                'monthly': dashboard_data['monthly'][mname],
                'weekly': dashboard_data['weekly'][mname],
            },
            labor_cost_data.get(mname)
        )
        anomaly_reason_data['monthly'][mname] = module_reason_data.get('monthly', {})
        anomaly_reason_data['weekly'][mname] = module_reason_data.get('weekly', {})
    print('  异常简析数据计算完成')

    # ============================================================
    # 结构分析数据计算
    # ============================================================
    print('计算结构分析数据...')

    # 展开周期所用的centers（需要拆分到需求部门的）
    EXPAND_CENTERS = {'运营中心', '创意中心'}

    structure_data = {'monthly': {}, 'weekly': {}}
    efficiency_analysis_data = {'monthly': {}, 'weekly': {}}

    for mod in MODULES:
        if should_hide_module(mod):
            continue
        module_options = build_module_runtime_options(mod)
        mname = module_options['name']
        structure_data['monthly'][mname] = {}
        structure_data['weekly'][mname] = {}
        efficiency_analysis_data['monthly'][mname] = {}
        efficiency_analysis_data['weekly'][mname] = {}

        for view in ['monthly', 'weekly']:
            periods = RECENT_MONTHS if view == 'monthly' else RECENT_WEEKS
            period_key = 'months' if view == 'monthly' else 'weeks'
            structure_data[view][mname] = build_structure_view_data(module_options, periods, period_key)
            if module_options.get('alternate_output_field'):
                structure_data[view][mname + '__handoff'] = build_structure_view_data(
                    module_options,
                    periods,
                    period_key,
                    module_options['alternate_output_field']
                )
            efficiency_analysis_data[view][mname] = build_efficiency_view_data(module_options, periods, period_key)

    print('  结构分析数据计算完成')

    # ============================================================
    # 人员效率分析数据
    # ============================================================
    print('计算人员效率分析数据...')
    person_efficiency_data = build_person_efficiency_data()
    print(f"  人员效率行: {len(person_efficiency_data['rows'])} 条")
    print(f"  人员明细行: {len(person_efficiency_data['details'])} 条")

    # ============================================================
    # 生成 HTML
    # ============================================================
    print('生成 HTML...')

    js_data = json.dumps(dashboard_data, ensure_ascii=False)
    labor_cost_js_data = json.dumps(labor_cost_data, ensure_ascii=False)
    anomaly_reason_js_data = json.dumps(anomaly_reason_data, ensure_ascii=False)
    structure_js_data = json.dumps(structure_data, ensure_ascii=False)
    efficiency_analysis_js_data = json.dumps(efficiency_analysis_data, ensure_ascii=False)
    person_efficiency_js_data = json.dumps(person_efficiency_data, ensure_ascii=False)
    month_labels = json.dumps(RECENT_MONTHS)
    labor_cost_month_labels = json.dumps(LABOR_COST_MONTHS)
    week_labels = json.dumps(RECENT_WEEKS)

    # 构建 ROI 数据（分部门分月）
    ROI_DEPT_ORDER = ['效果设计部','品牌设计部','合肥创意部','内容一部','内容二部']
    roi_months = sorted(set(str(row['年月']) for row in labor_cost_rows))
    roi_depts_all = set(str(row['制作部门']) for row in labor_cost_rows)
    roi_depts = [d for d in ROI_DEPT_ORDER if d in roi_depts_all]
    roi_data = {'months': roi_months, 'depts': roi_depts, 'values': {}, 'annual_roi': {}}
    # 计算各部门年度ROI = 当年总收入 / 当年总成本
    dept_income = {}
    dept_cost = {}
    for dept in roi_depts:
        dept_income[dept] = 0
        dept_cost[dept] = 0
        roi_data['values'][dept] = {}
        for row in labor_cost_rows:
            if str(row['制作部门']) == dept:
                roi_data['values'][dept][str(row['年月'])] = round(row['ROI'], 2)
                dept_income[dept] += (row.get('收入', 0) or 0)
                dept_cost[dept] += (row.get('成本', 0) or 0)
    for dept in roi_depts:
        roi_data['annual_roi'][dept] = round(dept_income[dept] / dept_cost[dept], 2) if dept_cost[dept] > 0 else 0
    roi_js_data = json.dumps(roi_data, ensure_ascii=False)

    with open(TEMPLATE_PATH, 'r', encoding='utf-8') as f:
        html_template = f.read()

    html = (html_template
        .replace('{{DATA_JSON}}', js_data)
        .replace('{{LABOR_COST_DATA_JSON}}', labor_cost_js_data)
        .replace('{{ANOMALY_REASON_DATA_JSON}}', anomaly_reason_js_data)
        .replace('{{STRUCTURE_DATA_JSON}}', structure_js_data)
        .replace('{{EFFICIENCY_ANALYSIS_DATA_JSON}}', efficiency_analysis_js_data)
        .replace('{{PERSON_EFFICIENCY_DATA_JSON}}', person_efficiency_js_data)
        .replace('{{MONTH_LABELS_JSON}}', month_labels)
        .replace('{{LABOR_COST_MONTH_LABELS_JSON}}', labor_cost_month_labels)
        .replace('{{WEEK_LABELS_JSON}}', week_labels)
        .replace('{{ROI_DATA_JSON}}', roi_js_data)
    )

    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write(html)

    print(f'  仪表盘已生成: {OUTPUT_PATH}')
    print('完成!')


if __name__ == '__main__':
    main()
